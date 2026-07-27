#!/usr/bin/env python3
"""Turn an estimate into a formula-live Excel workbook — the model, not a picture of one.

An estimate is a spreadsheet wearing prose. In markdown, changing one line item from 20 days to
26 means recomputing the total, the PERT mean, sigma, P50, P80, contingency, labour cost, the
cost base at two confidence levels, four margin figures and twelve mark scores — by hand, in a
document nobody can recalculate. That is how the arithmetic slips that this pack has already
caught twice get in, and it is why a reviewer cannot do the one thing a reviewer most wants to
do: change an input and see what happens.

So the workbook carries **formulas, not values**. Every derived cell is a real Excel formula over
input cells, and the input cells are the ones a reviewer is invited to move: rates, day counts,
the correlation assumption, the assumed competitor range. Nothing downstream has to be re-entered.

ONE MAINTAINED ARTEFACT. The workbook holds everything — the numbers as formulas AND the
judgement (basis of estimate, techniques and their reconciliation, the outside view, calibration,
contingency, the pricing-document mapping, re-baseline triggers) on their own sheets. The markdown
is a GENERATED SNAPSHOT of it, never a second thing to keep in step.

    workbook  (source of truth)  ──  --to-md  ─►  markdown snapshot (generated, read-only)

The markdown seeds the workbook ONCE. After that the direction reverses, and seeding again is a
destructive act that needs `--reseed` — the first version silently rebuilt the workbook from the
markdown on every run, so a reviewer's edits were wiped by the next build while the docs told
them the workbook was the model.

The snapshot is not decoration: it is what lets `eng_lint` read the estimate at all, and what
makes `git diff` show which numbers moved between two re-prices. A binary does neither.

Usage:
    python3 build_estimate_workbook.py <estimation.md>                  # seed the workbook once
    python3 build_estimate_workbook.py --out estimation.xlsx --to-md    # the normal direction
    python3 build_estimate_workbook.py <estimation.md> --reseed         # discard workbook edits
    python3 build_estimate_workbook.py --blank --out estimation.xlsx    # starter, no markdown yet
    python3 build_estimate_workbook.py <estimation.md> --check          # arithmetic audit only

Requires openpyxl (pip install openpyxl).
"""
from __future__ import annotations

import argparse
import math
from copy import copy
import os
import pathlib
import re
import sys

# Conventions are NOT invented here. They are the `xlsx` skill's financial-model conventions,
# which is the firm-wide standard every other spreadsheet already follows: blue text for
# hardcoded inputs, black for formulas, green for cross-sheet links, yellow fill for cells the
# reader is meant to edit, Arial throughout. A model that colour-codes itself differently from
# every other model in the building teaches the reader nothing and costs them a legend.
FONT = "Arial"
BLUE = "0000FF"            # hardcoded input / scenario lever
BLACK = "000000"           # formula
GREEN = "008000"           # formula referencing another sheet
YELLOW = "FFFF00"          # fill: edit this cell
HEAD_FILL = "1F3864"
EUR = '€#,##0'
PCT = '0.0%'               # stored as a fraction, per the xlsx skill
MULT = '0.0"x"'
# Legacy aliases so the per-cell calls below stay readable; `enforce_conventions` is what
# actually decides colour, from what each cell IS rather than from where it was written.
INPUT_FILL = YELLOW
DERIVED_FILL = "FFFFFF"


def _cells(line):
    """Split a markdown table row into stripped cells."""
    return [c.strip() for c in line.strip().strip("|").split("|")]


def _num(s):
    """A number out of a markdown cell, or None. Tolerates €, commas, bold, footnotes."""
    if s is None:
        return None
    s = re.sub(r"[*`€£$,\s]|\(.*?\)", "", str(s))
    s = s.replace("−", "-").replace("–", "-")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else None


TABLE_KINDS = ("effort", "overlap", "grades", "ratecard", "client", "scope", "certain",
               "schedule", "milestones")

# The judgement-bearing sections. These used to live only in the markdown, which made the
# markdown a second thing to maintain — and since the builder regenerated the workbook FROM it,
# a reviewer's edits in the workbook were silently overwritten on the next build. The workbook
# is the single source now, so these have to live in it too.
_MARKER = re.compile(r"^\s*<!--\s*table:(\w+)\s*-->\s*$")

#: The judgement topics. They used to be eight separate worksheets, which meant a workbook where
#: 8 of 21 tabs held four words each — and, worse, a tab called `BasisOfEstimate` existing made a
#: missing basis of estimate look present. One sheet, one row per topic, emptiness visible in the
#: row: the reader opens one tab and sees exactly which judgement is still owed.
JUDGEMENT_SHEET = "Judgement"
NARRATIVE_SHEETS = [
    ("BasisOfEstimate", ["1. Basis of estimate"]),
    ("Techniques", ["2. Techniques"]),
    ("OutsideView", ["2b. Outside view"]),
    ("Phasing", ["3c. By phase"]),
    ("Calibration", ["6. Calibration"]),
    ("Contingency", ["7. Contingency"]),
    ("PricingDoc", ["10. Into the buyer"]),
    ("Triggers", ["11. Re-baseline"]),
]


def parse_narrative(md):
    """Section body text (prose and any table rows), keyed by sheet name."""
    lines = md.splitlines()
    heads = [(i, l[3:].strip()) for i, l in enumerate(lines) if l.startswith("## ")]
    out = {}
    for sheet, prefixes in NARRATIVE_SHEETS:
        body = []
        # the template writes "## 1. Basis of estimate"; the workbook's own export writes
        # "## BasisOfEstimate". Accept both, or round-tripping quietly drops the judgement.
        accept = tuple(prefixes) + (sheet,)
        for n, (i, title) in enumerate(heads):
            if not any(title.startswith(pref) for pref in accept):
                continue
            end = heads[n + 1][0] if n + 1 < len(heads) else len(lines)
            for raw in lines[i + 1:end]:
                s = raw.rstrip()
                if _MARKER.match(s):
                    continue
                body.append(s)
        while body and not body[0].strip():
            body.pop(0)
        while body and not body[-1].strip():
            body.pop()
        out[sheet] = body
    return out


GENERATED_BANNER = "DO NOT EDIT THIS FILE"


def is_generated_snapshot(md):
    """Is this markdown the workbook's own export rather than a hand-written seed?

    Seeding reads `<!--table:effort-->` markers, which a snapshot does not carry — so seeding
    from one produces a workbook of placeholder rows and zero days, silently, and `--reseed`
    then overwrites the real model with it. Found by doing exactly that.
    """
    head = "\n".join(md.splitlines()[:12])
    return GENERATED_BANNER in head


def parse_tables(md):
    """Pull the model-bearing tables out of the estimation markdown, located by explicit markers.

    Each machine-read table is preceded by `<!--table:effort-->` and friends, planted by the
    template. The first version sniffed header text instead, and on the real GNI estimate it
    matched the §1 assumptions table as if it were the client-effort table — pulling "S-10, S-13
    (42 d)" out as `-10 sessions`. Nothing crashed; the client-hours total was simply wrong by a
    factor of seven. Guessing which table is which from its wording is not robust to an author
    rewording a heading, and the failure is silent, which is the worst combination.

    A marker with no table, or a table with no marker, yields nothing for that kind and is
    reported by the caller — never a silent zero, which reads like a real answer.
    """
    lines = md.splitlines()
    found = {k: [] for k in TABLE_KINDS}

    def rows_after(i):
        """Rows of the first markdown table starting at or after line i (skipping its header
        and the |---| separator). Stops at the first non-table line."""
        j = i
        while j < len(lines) and not lines[j].strip().startswith("|"):
            if lines[j].strip() and not _MARKER.match(lines[j]):
                return []                               # something other than blank intervened
            j += 1
        out = []
        for line in lines[j + 2:]:
            if not line.strip().startswith("|"):
                break
            out.append(_cells(line))
        return out

    def clean(s):
        return re.sub(r"[*`]", "", s).strip()

    for i, line in enumerate(lines):
        m = _MARKER.match(line)
        if not m or m.group(1) not in TABLE_KINDS:
            continue
        kind, rows = m.group(1), rows_after(i + 1)
        for r in rows:
            if kind == "effort" and len(r) >= 6:
                sid, o, mm, p = clean(r[0]), _num(r[3]), _num(r[4]), _num(r[5])
                if re.match(r"^S-\d+$", sid) and None not in (o, mm, p):
                    found["effort"].append((sid, clean(r[1]), o, mm, p,
                                            clean(r[-1]) if len(r) > 7 else ""))
            elif kind == "overlap" and len(r) >= 3:
                item, a, b = clean(r[0]), _num(r[1]), _num(r[2])
                if item and not item.upper().startswith("TOTAL") and None not in (a, b):
                    found["overlap"].append((item, a, b, clean(r[3]) if len(r) > 3 else ""))
            elif kind == "grades" and len(r) >= 3:
                g, rate, days = clean(r[0]), _num(r[1]), _num(r[2])
                if g and "blended" not in g.lower() and rate and days:
                    found["grades"].append((g, rate, days, clean(r[-1])))
            elif kind == "client" and len(r) >= 5:
                who, people, sess, hrs = clean(r[0]), _num(r[2]), _num(r[3]), _num(r[4])
                if who and not who.upper().startswith("TOTAL") and None not in (people, sess, hrs):
                    found["client"].append((who, clean(r[1]), people, sess, hrs,
                                            clean(r[6]) if len(r) > 6 else ""))
            elif kind == "scope" and len(r) >= 3:
                a, d = clean(r[0]), _num(r[2])
                if re.match(r"^A\d+$", a) and d:
                    found["scope"].append((a, clean(r[1]), d))
            elif kind == "ratecard" and len(r) >= 2:
                g, cost = clean(r[0]), _num(r[1])
                if g and cost is not None:
                    found["ratecard"].append((g, cost, _num(r[2]) if len(r) > 2 else None,
                                              clean(r[4]) if len(r) > 4 else ""))
            elif kind == "schedule" and len(r) >= 4:
                sid, people, util = clean(r[0]), _num(r[1]), _num(r[2])
                if re.match(r"^S-\d+$", sid) and people:
                    found["schedule"].append((sid, people, (util or 60) / 100 if (util or 0) > 1
                                              else (util or 0.6),
                                              clean(r[3]) if len(r) > 3 else "",
                                              _num(r[4]) if len(r) > 4 else 0,
                                              # the Span was parsed-then-dropped for months:
                                              # build() reads index 5, this appended five
                                              _num(r[5]) if len(r) > 5 else None))
            elif kind == "milestones" and len(r) >= 3:
                gate, wk = clean(r[0]), _num(r[1])
                if gate and wk:
                    found["milestones"].append((gate, wk, clean(r[2]),
                                                clean(r[3]) if len(r) > 3 else ""))
            elif kind == "certain" and len(r) >= 2:
                n, v = clean(r[0]), _num(r[1])
                # €0 is a real answer (the template's own first row is 'Financing | 0');
                # `and v` treated it as absent and understated non-labour cost silently
                if n and not n.upper().startswith("TOTAL") and v is not None:
                    found["certain"].append((n, v, clean(r[2]) if len(r) > 2 else ""))
    return found


# ── workbook construction ──────────────────────────────────────────────────────

def _style(ws, row, ncols, fill=None, bold=False, header=False):
    from openpyxl.styles import Font, PatternFill, Alignment
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        if header:
            cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=HEAD_FILL)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        else:
            if bold:
                cell.font = Font(name=FONT, bold=True, size=10)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)


def enforce_conventions(wb):
    """Apply the `xlsx` skill's model conventions across every sheet, in one pass.

    Threading font and colour through every write site is how one cell gets missed; doing it
    once at the end cannot miss any. Colour is derived from what the cell IS — a formula string
    means a formula, a cross-sheet reference means green — so it stays correct even when a cell
    changes kind later.
    """
    from openpyxl.styles import Font, PatternFill
    yellow = PatternFill("solid", fgColor=YELLOW)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                f = cell.font
                if f and f.color and f.color.rgb == "FFFFFFFF":
                    cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
                    continue                       # header band — leave it
                v = cell.value
                is_yellow = bool(cell.fill and cell.fill.fgColor
                                 and str(cell.fill.fgColor.rgb).endswith(YELLOW))
                if isinstance(v, str) and v.startswith("="):
                    colour = GREEN if "!" in v or any(
                        n in v for n in ("ratecard", "P50", "P80", "blended_rate", "rho",
                                         "nonlabour", "conf", "term_days", "low_", "max_marks",
                                         "sigma_cells", "pert_mean", "sum_", "costbase",
                                         "contingency_days", "scope_exposure", "rate_status")
                    ) else BLACK
                    cell.font = Font(name=FONT, size=10, bold=bool(f and f.bold), color=colour)
                elif is_yellow or (v is not None and not isinstance(v, str)):
                    # Yellow fill, or a typed literal, means an input — blue text either way.
                    cell.font = Font(name=FONT, size=10, bold=bool(f and f.bold), color=BLUE)
                    cell.fill = yellow
                else:
                    cell.font = Font(name=FONT, size=10, bold=bool(f and f.bold), color=BLACK)


def _widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _named(wb, name, sheet, ref):
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names.add(DefinedName(name, attr_text=f"'{sheet}'!{ref}"))


def build(data, rho, out_path, zero_rates=False, narrative=None):
    from openpyxl import Workbook

    wb = Workbook()
    narrative = narrative or {}
    eff = data["effort"] or [(f"S-{i:02d}", "<activity>", 0, 0, 0, "") for i in range(1, 4)]
    grades = data["grades"] or [("<grade>", 0, 0, "[⚠VERIFY]")]
    # A dedicated rate-card table wins; otherwise fall back to the rate column of the grades
    # table, so an estimate written before the rate card existed still produces a working model.
    ratecard = data.get("ratecard") or [(g, rate, None, src) for g, rate, _d, src in grades]
    client = data["client"] or [("<group>", "<activity>", 0, 0, 0, "")]
    scope = data["scope"] or [("A1", "<if wrong>", 0)]
    certain = data["certain"] or [("Financing", 0, ""), ("Travel", 0, ""), ("Production", 0, "")]

    # ── README ────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "README"
    _widths(ws, [110])
    for r, line in enumerate([
        "ESTIMATION MODEL — live formulas",
        "",
        "Yellow cells with blue text are INPUTS. Black formulas calculate locally; green formulas link sheets.",
        "",
        "Sheet            What it does",
        "Inputs           Rates, non-labour, correlation assumption, competitor range. Start here.",
        "Effort           One row per scope item (S-ID) with three-point O/M/P. Expected + sigma are formulas.",
        "OverlapAudit     Record of days removed for double-counting. Cross-checks against Effort.",
        "Range            P50 / P80 from the three-point data. The CORRELATION cell drives the width — see Inputs.",
        "Grades           Grade x rate x days -> labour cost and the blended rate.",
        "ClientEffort     The client's hours. Required by some RFPs; scope protection either way.",
        "ScopeVariance    What each assumption costs if it is wrong. A DIFFERENT quantity from the P50/P80 band.",
        "Schedule         Duration, dependencies and dates per activity. Effort / (people x util) — a",
        "                 cost model with no schedule cannot say when anything lands or who is free.",
        "Gantt            The picture of Schedule: week grid, gates, and the weekly FTE demand curve.",
        "Cost             Cost base at P50 and at P80. Stops before margin — pricing is a human decision.",
        "Decision         Candidate prices -> marks and margin. Deliberately recommends nothing.",
        "Judgement        The reasoning the numbers rest on — basis of estimate, techniques and their",
        "                 reconciliation, outside view, phasing, calibration, contingency, the pricing-",
        "                 document mapping, re-baseline triggers. One row per topic; blank rows say OWED.",
        "",
        "TWO UNCERTAINTIES, KEPT APART:",
        "  Range         = execution variance  (how long the work takes)   -> handled with contingency days",
        "  ScopeVariance = scope variance      (whether this IS the work)  -> handled with assumptions in the contract",
        "  Do not add them together. They are managed with different instruments.",
        "",
        "The correlation cell on Inputs matters more than it looks. Textbook PERT assumes line items",
        "are independent; on a decomposed consulting scope they share assumptions and a team, so when",
        "one is wrong several are wrong. If sigma looks implausibly tight, that cell is the reason.",
        "",
        "This workbook does not nominate a bid price. It supplies the tradeoff; the partner decides.",
    ], 1):
        ws.cell(row=r, column=1, value=line)
    ws.cell(row=1, column=1).font = __import__("openpyxl").styles.Font(bold=True, size=13)

    # ── Inputs ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Inputs")
    _widths(ws, [42, 16, 62])
    ws.append(["Input", "Value", "Note"]); _style(ws, 1, 3, header=True)
    inputs = [
        ("Working days in term", 126, "After public holidays. Drives implied FTE.", "term_days"),
        ("Correlation assumption (rho)", rho,
         "0 = independent (textbook PERT, usually too tight) · 0.5 = partially correlated "
         "(professional-services default) · 1 = fully correlated", "rho"),
        ("Confidence level for the commitment number", 0.80,
         "0.80 = P80. Change to 0.70 or 0.90 to re-cut the range.", "conf"),
        ("Assumed lowest competitor bid — low", 225000, "An assumption. Say so in the narrative.", "low_a"),
        ("Assumed lowest competitor bid — mid", 240000, "", "low_b"),
        ("Assumed lowest competitor bid — high", 270000, "", "low_c"),
        ("Cost marks available", 300, "From the buyer's evaluation table.", "max_marks"),
        ("Term (weeks)", 26, "Contract duration. The schedule may not end after it.", "term_weeks"),
        ("Working days per week", 5, "Duration = effort / (people x this x utilisation).", "days_per_week"),
        ("Project start date", __import__("datetime").date.today(),
         "Week 1 Monday. Every date on Schedule is derived from it.", "start_date"),
    ]
    # Row numbers are tracked explicitly rather than inferred from len(). Mixing ws.append()
    # with ws.cell(row=N) moves openpyxl's internal cursor in a way that is easy to get wrong,
    # and the first version did: the `nonlabour` named range landed one row short of the SUM,
    # so the cost base silently omitted €11k. The workbook still opened and still looked right.
    row = 1
    for label, val, note, nm in inputs:
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=3, value=note)
        _style(ws, row, 2, fill=INPUT_FILL)
        if nm == "start_date":
            ws.cell(row=row, column=2).number_format = "dd-mmm-yy"
        _named(wb, nm, "Inputs", f"$B${row}")
    row += 2
    ws.cell(row=row, column=1, value="Non-labour and certain costs").font = \
        __import__("openpyxl").styles.Font(bold=True)
    row += 1
    for c, h in enumerate(("Line", "€", "Basis"), 1):
        ws.cell(row=row, column=c, value=h)
    _style(ws, row, 3, header=True)
    cert_first = row + 1
    for name, val, basis in certain:
        row += 1
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=3, value=basis)
        _style(ws, row, 2, fill=INPUT_FILL)
    cert_last = row
    row += 1
    ws.cell(row=row, column=1, value="Total non-labour + certain")
    ws.cell(row=row, column=2, value=f"=SUM(B{cert_first}:B{cert_last})")
    _style(ws, row, 2, fill=DERIVED_FILL, bold=True)
    ws.cell(row=row, column=2).number_format = EUR
    _named(wb, "nonlabour", "Inputs", f"$B${row}")

    # ── RateCard ──────────────────────────────────────────────────────────────
    # One table, one place. Rates typed inline against each grade row mean the day the real card
    # arrives you edit every row and hope you caught them all; here you paste one table and the
    # whole model recalculates. It doubles as the deliverable some tenders require in their own
    # right (an extension / call-off rate schedule), which is why sell rates sit alongside cost.
    ws = wb.create_sheet("RateCard")
    _widths(ws, [44, 18, 18, 16, 52])
    ws.cell(row=1, column=1, value="RATE CARD — the single source of every rate in this model")
    ws.cell(row=1, column=1).font = __import__("openpyxl").styles.Font(bold=True, size=12)
    ws.cell(row=2, column=1, value="Status")
    ws.cell(row=2, column=2, value="PLACEHOLDER")
    _style(ws, 2, 2, fill=INPUT_FILL, bold=True)
    ws.cell(row=2, column=3,
            value="Set to ACTUAL only when these are the firm's real published rates. "
                  "While it reads PLACEHOLDER, no price computed here may be committed.")
    _named(wb, "rate_status", "RateCard", "$B$2")
    for c, h in enumerate(("Grade", "Cost rate / day €", "Sell rate / day €",
                           "Effective date", "Source"), 1):
        ws.cell(row=4, column=c, value=h)
    _style(ws, 4, 5, header=True)
    rc_first = 5
    for i, (g, cost, sell, src) in enumerate(ratecard, rc_first):
        ws.cell(row=i, column=1, value=g)
        ws.cell(row=i, column=2, value=0 if zero_rates else cost)
        ws.cell(row=i, column=3, value=None if zero_rates else sell)
        ws.cell(row=i, column=4, value=None)
        ws.cell(row=i, column=5, value=src)
        _style(ws, i, 4, fill=INPUT_FILL)
        ws.cell(row=i, column=2).number_format = EUR
        ws.cell(row=i, column=3).number_format = EUR
    rc_last = rc_first + len(ratecard) - 1
    _named(wb, "ratecard", "RateCard", f"$A${rc_first}:$C${rc_last}")
    _named(wb, "ratecard_grades", "RateCard", f"$A${rc_first}:$A${rc_last}")
    for k, line in enumerate([
        "",
        "Grade names must match the Grades sheet exactly — that sheet looks rates up from here.",
        "A grade with no row here resolves to a rate of 0, and Grades flags it. Zero is a",
        "visible wrong answer; a lookup error spreading through the model is not.",
        "",
        "Sell rates are not used by the cost model (which runs on cost rates). They are here",
        "because a tender often has to submit a rate schedule for extension or call-off work,",
        "and that schedule should not be maintained in a second place.",
    ]):
        ws.cell(row=rc_last + 2 + k, column=1, value=line)

    # ── Effort ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Effort")
    _widths(ws, [10, 46, 8, 8, 8, 12, 10, 24])
    ws.append(["S-ID", "Activity", "O", "M", "P", "Expected", "sigma", "Lead role"])
    _style(ws, 1, 8, header=True)
    for i, (sid, act, o, m, p, lead) in enumerate(eff, 2):
        ws.append([sid, act, o, m, p,
                   f"=(C{i}+4*D{i}+E{i})/6", f"=(E{i}-C{i})/6", lead])
        for col in ("C", "D", "E"):
            ws[f"{col}{i}"].fill = __import__("openpyxl").styles.PatternFill("solid", fgColor=INPUT_FILL)
        for col in ("F", "G"):
            ws[f"{col}{i}"].fill = __import__("openpyxl").styles.PatternFill("solid", fgColor=DERIVED_FILL)
            ws[f"{col}{i}"].number_format = "0.0"
    first, last = 2, len(eff) + 1
    tot = last + 1
    ws.append(["TOTAL", "", f"=SUM(C{first}:C{last})", f"=SUM(D{first}:D{last})",
               f"=SUM(E{first}:E{last})", f"=SUM(F{first}:F{last})",
               f"=SUM(G{first}:G{last})", ""])
    _style(ws, tot, 8, fill=DERIVED_FILL, bold=True)
    for nm, ref in (("sum_O", f"$C${tot}"), ("sum_M", f"$D${tot}"), ("sum_P", f"$E${tot}"),
                    ("pert_mean", f"$F${tot}")):
        _named(wb, nm, "Effort", ref)
    _named(wb, "sigma_cells", "Effort", f"$G${first}:$G${last}")
    ws.append([])
    ws.append(["Completeness sweep — booked somewhere above?", "", "", "", "", "", "", ""])
    _style(ws, tot + 2, 1, bold=True)
    for k, item in enumerate(["mobilisation", "governance / steering", "QA",
                              "document production", "client review cycles + rework"]):
        ws.append([""] + [item, "yes / no"])
        _style(ws, tot + 3 + k, 3, fill=INPUT_FILL)

    # ── OverlapAudit ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("OverlapAudit")
    _widths(ws, [34, 12, 12, 12, 34, 56])
    ws.append(["Item", "Original", "Revised", "Deduction", "Pattern",
               "Why it re-uses another item's fieldwork"])
    _style(ws, 1, 6, header=True)
    ov = data["overlap"]
    for i, (item, a, b, why) in enumerate(ov or [("<item>", 0, 0, "")], 2):
        ws.append([item, a, b, f"=B{i}-C{i}", "", why])
        _style(ws, i, 3, fill=INPUT_FILL)
        ws[f"D{i}"].fill = __import__("openpyxl").styles.PatternFill("solid", fgColor=DERIVED_FILL)
    olast = max(2, len(ov) + 1)
    ws.append(["TOTAL DEDUCTION", f"=SUM(B2:B{olast})", f"=SUM(C2:C{olast})",
               f"=SUM(D2:D{olast})", "", ""])
    _style(ws, olast + 1, 6, fill=DERIVED_FILL, bold=True)
    ws.append([])
    ws.append(["Naive bottom-up (before audit)", f"=pert_mean+D{olast + 1}", "", "",
               "", "Cross-check: audited Effort total + deduction"])
    ws.append(["Audited (Effort!TOTAL Expected)", "=pert_mean", "", "", "", ""])
    for rr in (olast + 3, olast + 4):
        _style(ws, rr, 2, fill=DERIVED_FILL)
    ws.append([])
    ws.append(["Examined and KEPT (an audit that only ever cuts is a discount, not an audit)"])
    _style(ws, olast + 6, 1, bold=True)
    ws.append(["<item — genuinely separate fieldwork because …>"])

    # ── Range ─────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Range")
    _widths(ws, [46, 16, 76])
    ws.append(["Execution variance", "Days", "Note"]); _style(ws, 1, 3, header=True)
    rows = [
        ("sigma if INDEPENDENT (textbook PERT)", "=SQRT(SUMSQ(sigma_cells))",
         "Assumes line items are unrelated. On a decomposed consulting scope they are not."),
        ("sigma if FULLY CORRELATED", "=SUM(sigma_cells)",
         "Every item hangs on the same handful of assumptions."),
        ("sigma USED (rho from Inputs)", "=IFERROR(SQRT(B2^2+rho*(B3^2-B2^2)),0)",
         "Change rho on Inputs. If sigma looks implausibly tight, rho is the reason."),
        ("z for the confidence level", "=NORMSINV(conf)",
         "0.8416 at P80. Replace with a constant if your spreadsheet lacks NORMSINV."),
        ("Sum O (everything goes right)", "=sum_O", ""),
        ("Sum M (most likely)", "=sum_M", ""),
        ("Sum P (everything goes wrong)", "=sum_P", ""),
        ("P50 — the planning number", "=pert_mean", "PERT mean."),
        ("P80 — the commitment number", "=B9+B5*B4", "P50 + z * sigma."),
        ("Contingency (P80 - P50)", "=B10-B9", "This IS the contingency. Do not also hand-pick risk lines."),
        ("Implied FTE at P50", "=IFERROR(B9/term_days,0)", "Compare against the team actually being named."),
    ]
    for i, (label, f, note) in enumerate(rows, 2):
        ws.append([label, f, note])
        _style(ws, i, 2, fill=DERIVED_FILL)
        ws.cell(row=i, column=2).number_format = "0.0"
    for nm, ref in (("P50", "$B$9"), ("P80", "$B$10"), ("contingency_days", "$B$11")):
        _named(wb, nm, "Range", ref)
    _style(ws, 9, 3, bold=True); _style(ws, 10, 3, bold=True)

    # ── Schedule ──────────────────────────────────────────────────────────────
    # An estimate with a cost base and no schedule cannot answer the two questions a buyer asks
    # next: when does it land, and who is on it in week 9. Duration is NOT effort — it is
    # effort / (people x working days x utilisation), which is why the same 240 days can be six
    # months or eighteen. So the schedule is DERIVED from the same Effort rows the cost is
    # derived from: move a day count and the duration, the dates, the FTE curve and the term
    # check all move with it.
    #
    # Two kinds of activity, because conflating them is the classic schedule error:
    #   DISCRETE       — duration falls out of effort and staffing (leave Span blank)
    #   LEVEL OF EFFORT — governance, QA, review cycles: the span is given by the term and the
    #                     STAFFING falls out of it instead (put the span in, in weeks)
    ws = wb.create_sheet("Schedule")
    _widths(ws, [9, 38, 9, 8, 7, 8, 10, 11, 8, 8, 8, 11, 11, 30, 22])
    ws.append(["S-ID", "Activity", "Effort d", "People", "Util", "Span", "Duration wk",
               "Predecessor", "Start / lag", "Start wk", "End wk", "Start date", "End date",
               "Bar (wk 1 →)", "Lead role"])
    _style(ws, 1, 15, header=True)
    sched_in = {row[0]: row for row in (data.get("schedule") or [])}
    n_eff = len(eff)
    for i, (sid, act, _o, _m, _p, lead) in enumerate(eff, 2):
        people, util, pred, lag, span = 2, 0.6, "", 1, None
        if sid in sched_in:
            _s, people, util, pred, lag = sched_in[sid][:5]
            span = sched_in[sid][5] if len(sched_in[sid]) > 5 else None
        ws.append([
            sid, act,
            f"=ROUND(Effort!F{i},1)",                      # green: one effort number, two models
            people, util, span,
            # a given span wins; otherwise duration falls out of effort and staffing
            f"=IF(F{i}>0,F{i},IFERROR(MAX(1,CEILING(C{i}/(D{i}*days_per_week*E{i}),1)),1))",
            pred, lag or 1,
            # no predecessor → the cell IS the start week; with one → finish-to-start, where a
            # blank or zero lag means the following week (0 would overlap the predecessor's last)
            f'=IF($H{i}="",MAX(1,$I{i}),IFERROR(INDEX($K$2:$K${n_eff + 1},'
            f'MATCH($H{i},$A$2:$A${n_eff + 1},0))+MAX(1,$I{i}),1))',
            f"=J{i}+G{i}-1",
            f"=start_date+(J{i}-1)*7", f"=start_date+K{i}*7-1",
            f'=REPT("·",J{i}-1)&REPT("█",G{i})',
            lead,
        ])
        for col in ("D", "E", "F", "H", "I"):              # the levers a reviewer moves
            ws[f"{col}{i}"].fill = __import__("openpyxl").styles.PatternFill("solid", fgColor=INPUT_FILL)
        ws[f"E{i}"].number_format = PCT
        ws[f"C{i}"].number_format = "0.0"
        for col in ("L", "M"):
            ws[f"{col}{i}"].number_format = "dd-mmm-yy"
    slast = n_eff + 1
    srow = slast + 1
    ws.append(["TOTAL / SPAN", "", f"=ROUND(SUM(C2:C{slast}),1)", "", "", "",
               f"=MAX(K2:K{slast})-MIN(J2:J{slast})+1", "", "",
               f"=MIN(J2:J{slast})", f"=MAX(K2:K{slast})",
               "=start_date", f"=start_date+MAX(K2:K{slast})*7-1", "", ""])
    _style(ws, srow, 15, fill=DERIVED_FILL, bold=True)
    for col in ("L", "M"):
        ws[f"{col}{srow}"].number_format = "dd-mmm-yy"
    _named(wb, "sched_start", "Schedule", f"$J$2:$J${slast}")
    _named(wb, "sched_end", "Schedule", f"$K$2:$K${slast}")
    _named(wb, "sched_days", "Schedule", f"$C$2:$C${slast}")
    _named(wb, "sched_dur", "Schedule", f"$G$2:$G${slast}")
    _named(wb, "last_week", "Schedule", f"$K${srow}")

    # The three checks that make a schedule worth trusting. Each says what a failure MEANS,
    # because "FALSE" in a cell teaches nobody anything.
    ws.append([])
    ws.append(["CHECKS"])
    _style(ws, srow + 2, 1, bold=True)
    checks = [
        ("Every estimated day is scheduled",
         f'=IF(ABS(C{srow}-ROUND(pert_mean,1))<0.2,"OK","MISMATCH: schedule covers "'
         f'&TEXT(C{srow},"0.0")&" d against an estimate of "&TEXT(pert_mean,"0.0")&" d")',
         "Schedule and cost must come from the same effort rows, or the plan prices work it does not contain."),
        ("Schedule fits the term",
         '=IF(last_week<=term_weeks,"OK","OVERRUN: ends week "&last_week&" against a "'
         '&term_weeks&"-week term")',
         "A plan that ends after the contract is a commitment nobody can make."),
        ("Peak staffing is deliverable",
         '=TEXT(MAX(Gantt!$B$4:$AA$4),"0.0")&" FTE peak vs "&TEXT(Range!$B$12,"0.0")'
         '&" average — the named team has to cover the peak, not the average"',
         "Averaging a 26-week engagement hides the week nobody is available."),
    ]
    for k, (label, formula, why) in enumerate(checks):
        rr = srow + 3 + k
        ws.cell(row=rr, column=1, value=label)
        ws.cell(row=rr, column=3, value=formula)
        ws.cell(row=rr, column=7, value=why)
        _style(ws, rr, 3, fill=DERIVED_FILL)

    # Milestones: the dates a buyer reads, and what each one decides.
    mrow = srow + 3 + len(checks) + 1
    ws.cell(row=mrow, column=1, value="MILESTONES / GATES").font = \
        __import__("openpyxl").styles.Font(bold=True, name=FONT)
    mrow += 1
    for c, h in enumerate(("Gate", "Week", "Date", "What it decides", "Deliverable"), 1):
        ws.cell(row=mrow, column=c, value=h)
    _style(ws, mrow, 5, header=True)
    miles = data.get("milestones") or [
        ("G1", 12, "current-state baseline accepted", "D1"),
        ("G2", 18, "options decision taken", "D2"),
        ("G3", 24, "recommendation, roadmap and cost accepted", "D3-D5"),
        ("G4", 26, "final acceptance", "D6"),
    ]
    mfirst = mrow + 1
    for k, (gate, wk, decides, deliv) in enumerate(miles):
        rr = mfirst + k
        ws.cell(row=rr, column=1, value=gate)
        ws.cell(row=rr, column=2, value=wk)
        ws.cell(row=rr, column=3, value=f"=start_date+B{rr}*7-1")
        ws.cell(row=rr, column=4, value=decides)
        ws.cell(row=rr, column=5, value=deliv)
        _style(ws, rr, 2, fill=INPUT_FILL)
        ws.cell(row=rr, column=3).number_format = "dd-mmm-yy"
    _named(wb, "gate_weeks", "Schedule", f"$B${mfirst}:$B${mfirst + len(miles) - 1}")
    _named(wb, "gate_names", "Schedule", f"$A${mfirst}:$A${mfirst + len(miles) - 1}")
    ws.freeze_panes = "C2"

    # ── Gantt ─────────────────────────────────────────────────────────────────
    # The picture, kept apart from the model: one column per week, every cell a formula over
    # Schedule. Conditional formatting paints the bars, so nothing is hand-drawn and nothing
    # goes stale when a duration changes.
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("Gantt")
    weeks = 26
    _widths(ws, [34] + [3.2] * weeks)
    ws.cell(row=1, column=1, value="Week →")
    for w in range(1, weeks + 1):
        c = ws.cell(row=1, column=1 + w, value=w)
        c.alignment = __import__("openpyxl").styles.Alignment(horizontal="center")
    _style(ws, 1, weeks + 1, header=True)
    ws.cell(row=2, column=1, value="Gate")
    for w in range(1, weeks + 1):
        col = get_column_letter(1 + w)
        ws.cell(row=2, column=1 + w,
                value=f'=IFERROR(INDEX(gate_names,MATCH({col}$1,gate_weeks,0)),"")')
        ws.cell(row=2, column=1 + w).alignment = \
            __import__("openpyxl").styles.Alignment(horizontal="center")
    ws.cell(row=3, column=1, value="FTE demand")
    ws.cell(row=4, column=1, value="  (people needed that week)")
    for w in range(1, weeks + 1):
        col = get_column_letter(1 + w)
        ws.cell(row=4, column=1 + w, value=(
            f"=IFERROR(ROUND(SUMPRODUCT((sched_start<={col}$1)*(sched_end>={col}$1)"
            f"*(sched_days/sched_dur/days_per_week)),1),0)"))
        ws.cell(row=4, column=1 + w).number_format = "0.0"
    grow = 6
    ws.cell(row=grow - 1, column=1, value="Activity").font = \
        __import__("openpyxl").styles.Font(bold=True, name=FONT)
    for i in range(2, slast + 1):
        rr = grow + i - 2
        ws.cell(row=rr, column=1, value=f"=Schedule!A{i}&\"  \"&Schedule!B{i}")
        for w in range(1, weeks + 1):
            col = get_column_letter(1 + w)
            ws.cell(row=rr, column=1 + w, value=(
                f"=IF(AND({col}$1>=Schedule!$J{i},{col}$1<=Schedule!$K{i}),1,\"\")"))
    glast = grow + slast - 2
    bars = f"B{grow}:{get_column_letter(1 + weeks)}{glast}"
    ws.conditional_formatting.add(bars, CellIsRule(
        operator="equal", formula=["1"],
        fill=__import__("openpyxl").styles.PatternFill("solid", fgColor="2A2D9C"),
        font=__import__("openpyxl").styles.Font(color="2A2D9C", name=FONT)))
    ws.conditional_formatting.add(
        f"B4:{get_column_letter(1 + weeks)}4",
        CellIsRule(operator="greaterThan", formula=["Range!$B$12"],
                   fill=__import__("openpyxl").styles.PatternFill("solid", fgColor="FFE0B2")))
    ws.cell(row=glast + 2, column=1,
            value="Bars are formulas over Schedule — change a duration there and this redraws.")
    ws.cell(row=glast + 3, column=1,
            value="Shaded FTE cells are weeks above the P50 average: that is where staffing bites.")
    ws.freeze_panes = "B6"

    # ── Grades ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Grades")
    _widths(ws, [44, 18, 10, 16, 30])
    for c, h in enumerate(("Grade", "Cost rate / day (from RateCard)", "Days", "Cost", "Check"), 1):
        ws.cell(row=1, column=c, value=h)
    _style(ws, 1, 5, header=True)
    for i, (g, _rate, days, _src) in enumerate(grades, 2):
        ws.cell(row=i, column=1, value=g)
        ws.cell(row=i, column=2, value=f"=IFERROR(VLOOKUP(A{i},ratecard,2,FALSE),0)")
        ws.cell(row=i, column=3, value=days)
        ws.cell(row=i, column=4, value=f"=B{i}*C{i}")
        ws.cell(row=i, column=5,
                value=f'=IF(COUNTIF(ratecard_grades,A{i})=0,"NOT IN RATE CARD","")')
        _style(ws, i, 1, fill=INPUT_FILL)          # grade name is an input …
        _style(ws, i, 3, fill=INPUT_FILL)          # … so are days
        ws[f"C{i}"].fill = __import__("openpyxl").styles.PatternFill("solid", fgColor=INPUT_FILL)
        for c in ("B", "D", "E"):
            ws[f"{c}{i}"].fill = __import__("openpyxl").styles.PatternFill("solid", fgColor=DERIVED_FILL)
        ws[f"B{i}"].number_format = '#,##0'
        ws[f"D{i}"].number_format = EUR
    glast = len(grades) + 1
    tot = glast + 1
    ws.cell(row=tot, column=1, value="BLENDED / TOTAL")
    # IFERROR, not a bare division: with a zeroed rate card or an empty grade list this must
    # return 0, not spray #DIV/0! through every sheet that depends on the blended rate.
    ws.cell(row=tot, column=2, value=f"=IFERROR(D{tot}/C{tot},0)")
    ws.cell(row=tot, column=3, value=f"=SUM(C2:C{glast})")
    ws.cell(row=tot, column=4, value=f"=SUM(D2:D{glast})")
    _style(ws, tot, 5, fill=DERIVED_FILL, bold=True)
    ws[f"B{tot}"].number_format = EUR
    ws[f"D{tot}"].number_format = EUR
    _named(wb, "blended_rate", "Grades", f"$B${tot}")
    _named(wb, "grade_days", "Grades", f"$C${tot}")
    ws.cell(row=tot + 2, column=1, value="Allocated days must equal P50")
    ws.cell(row=tot + 2, column=2, value=f'=IF(ABS(C{tot}-P50)<0.5,"OK","MISMATCH")')
    ws.cell(row=tot + 2, column=5, value="Re-allocate the grade mix when P50 moves.")
    _style(ws, tot + 2, 2, fill=DERIVED_FILL, bold=True)
    ws.cell(row=tot + 3, column=1, value="Rate card status")
    ws.cell(row=tot + 3, column=2, value="=rate_status")
    ws.cell(row=tot + 3, column=5, value="ACTUAL required before any price is committed.")
    _style(ws, tot + 3, 2, fill=DERIVED_FILL, bold=True)

    # ── ClientEffort ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("ClientEffort")
    _widths(ws, [40, 42, 10, 11, 13, 13, 16])
    ws.append(["Their role / group", "Activity", "People", "Sessions", "Hours each",
               "Total hours", "When"])
    _style(ws, 1, 7, header=True)
    for i, (who, act, people, sess, hrs, when) in enumerate(client, 2):
        ws.append([who, act, people, sess, hrs, f"=C{i}*D{i}*E{i}", when])
        _style(ws, i, 5, fill=INPUT_FILL)
        ws[f"F{i}"].fill = __import__("openpyxl").styles.PatternFill("solid", fgColor=DERIVED_FILL)
    clast = len(client) + 1
    ws.append(["TOTAL", "", "", "", "", f"=SUM(F2:F{clast})", ""])
    _style(ws, clast + 1, 7, fill=DERIVED_FILL, bold=True)
    ws.append(["Client days (hrs / 7.5)", "", "", "", "", f"=F{clast + 1}/7.5", ""])
    ws.append(["Client-to-consultant ratio", "", "", "", "", f"=IFERROR(F{clast + 2}/P50,0)",
               "10–15% is normal for a workshop-led assessment"])
    ws[f"F{clast + 3}"].number_format = "0.0%"
    for rr in (clast + 2, clast + 3):
        _style(ws, rr, 6, fill=DERIVED_FILL)

    # ── ScopeVariance ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("ScopeVariance")
    _widths(ws, [12, 62, 12, 16, 40])
    ws.append(["Assumption", "If wrong", "Days", "€", "Disposition"])
    _style(ws, 1, 5, header=True)
    for i, (a, why, d) in enumerate(scope, 2):
        ws.append([a, why, d, f"=C{i}*blended_rate", "excluded by assumption / accepted"])
        _style(ws, i, 3, fill=INPUT_FILL)
        ws[f"D{i}"].fill = __import__("openpyxl").styles.PatternFill("solid", fgColor=DERIVED_FILL)
        ws[f"D{i}"].number_format = EUR
    slast = len(scope) + 1
    ws.append(["TOTAL", "Upside exposure — carried by ASSUMPTIONS, not by days",
               f"=SUM(C2:C{slast})", f"=SUM(D2:D{slast})", ""])
    _style(ws, slast + 1, 5, fill=DERIVED_FILL, bold=True)
    _named(wb, "scope_exposure_eur", "ScopeVariance", f"$D${slast + 1}")
    ws.append([])
    ws.append(["Scope band ÷ execution band", f"=IFERROR(C{slast + 1}/contingency_days,0)", "",
               "", "If this is >1, scope definition is the dominant risk — not execution."])
    _style(ws, slast + 3, 2, fill=DERIVED_FILL, bold=True)
    ws.cell(row=slast + 3, column=2).number_format = MULT

    # ── Cost ──────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Cost")
    _widths(ws, [42, 18, 18, 56])
    ws.append(["Line", "at P50 €", "at P80 €", "Note"]); _style(ws, 1, 4, header=True)
    ws.append(["Labour", "=P50*blended_rate", "=P80*blended_rate", "Days x blended cost rate"])
    ws.append(["Non-labour + certain costs", "=nonlabour", "=nonlabour", "From Inputs"])
    ws.append(["COST BASE", "=B2+B3", "=C2+C3", "Margin is NOT set here — pricing is a human decision"])
    ws.append(["Rate card status", "=rate_status", "",
               '=IF(rate_status<>"ACTUAL","⚠ PLACEHOLDER RATES — this cost base is not committable","")'])
    for rr in (2, 3, 4):
        _style(ws, rr, 3, fill=DERIVED_FILL)
        for c in ("B", "C"):
            ws[f"{c}{rr}"].number_format = EUR
    _style(ws, 4, 4, bold=True)
    _named(wb, "costbase_P50", "Cost", "$B$4")
    _named(wb, "costbase_P80", "Cost", "$C$4")

    # ── Decision ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Decision")
    _widths(ws, [18, 18, 18, 18, 14, 18, 30])
    ws.append(["Candidate price", "Marks if lowest = low", "if mid", "if high",
               "Margin vs P50", "Margin €", "Flag"])
    _style(ws, 1, 7, header=True)
    for i in range(2, 8):
        # IFERROR as well as the blank guard: the blank test stops an empty row, but a reviewer
        # typing 0 into a candidate-price cell would still divide by zero and spray #DIV/0!
        # across the row they are in the middle of filling in.
        ws.append([
            None,
            f"=IFERROR(IF(A{i}=\"\",\"\",MIN(low_a/A{i},1)*max_marks),\"\")",
            f"=IFERROR(IF(A{i}=\"\",\"\",MIN(low_b/A{i},1)*max_marks),\"\")",
            f"=IFERROR(IF(A{i}=\"\",\"\",MIN(low_c/A{i},1)*max_marks),\"\")",
            f"=IFERROR(IF(A{i}=\"\",\"\",(A{i}-costbase_P50)/A{i}),\"\")",
            f"=IF(A{i}=\"\",\"\",A{i}-costbase_P50)",
            f'=IF(A{i}="","",IF(A{i}<costbase_P50,"UNDER COST BASE",'
            f'IF(A{i}<costbase_P80,"contingency not covered",'
            f'IF(A{i}-costbase_P50<scope_exposure_eur,"margin < excluded scope exposure","")))) ',
        ])
        _style(ws, i, 1, fill=INPUT_FILL)
        for c in ("B", "C", "D", "E", "F", "G"):
            ws[f"{c}{i}"].fill = __import__("openpyxl").styles.PatternFill("solid", fgColor=DERIVED_FILL)
        for c in ("B", "C", "D"):
            ws[f"{c}{i}"].number_format = "0.0"
        ws[f"E{i}"].number_format = PCT
        ws[f"F{i}"].number_format = EUR
    ws.append([])
    for k, line in enumerate([
        "Type candidate prices into column A. Everything else is a formula.",
        "Marks formula caps the ratio at 1: below the lowest bid you score full marks, not more.",
        "",
        "THIS SHEET DELIBERATELY RECOMMENDS NOTHING. Pricing weighs client appetite, competitive",
        "position and risk tolerance that this model does not contain. Record the decision, once",
        "taken, in _pm/raid_and_decisions.md.",
    ]):
        ws.cell(row=9 + k, column=1, value=line)
    ws.cell(row=12, column=1).font = __import__("openpyxl").styles.Font(bold=True)

    # ── Judgement ─────────────────────────────────────────────────────────────
    # The judgement the numbers rest on, in the same file as the numbers — one row per topic.
    # As eight separate tabs this was 8 of 21 worksheets holding four words each, and an empty
    # tab named `BasisOfEstimate` reads as "there is one" rather than "this is owed".
    ws = wb.create_sheet(JUDGEMENT_SHEET)
    _widths(ws, [26, 10, 120])
    ws.append(["Topic", "State", "What it says"])
    _style(ws, 1, 3, header=True)
    from openpyxl.styles import Alignment as _Al
    r = 1
    for sheet, _prefixes in NARRATIVE_SHEETS:
        body = [ln for ln in (narrative.get(sheet) or [])
                if ln.strip() and not (ln.strip().startswith("|") and set(ln.strip()) <= set("|-: "))]
        text = "\n".join(re.sub(r"[*`]", "", ln).strip() for ln in body).strip()
        r += 1
        ws.cell(row=r, column=1, value=sheet)
        ws.cell(row=r, column=2, value=f'=IF(LEN(C{r})<20,"OWED","written")')
        ws.cell(row=r, column=3, value=text or "<not yet written>")
        ws.cell(row=r, column=3).alignment = _Al(wrap_text=True, vertical="top")
        ws.cell(row=r, column=1).alignment = _Al(vertical="top")
        ws.row_dimensions[r].height = 58 if text else 15
    last = r
    ws.append([])
    ws.append(["JUDGEMENT OWED",
               f'=COUNTIF(B2:B{last},"OWED")&" of {len(NARRATIVE_SHEETS)} topics unwritten"',
               "A cost base whose basis, calibration and contingency rationale are blank is a "
               "number without an argument. Write the row, not a separate document."])
    _style(ws, last + 2, 2, fill=DERIVED_FILL, bold=True)

    # Tab order follows the model's flow, so a reviewer reads it in the order it computes.
    order_sheets(wb)

    enforce_conventions(wb)
    present_schedule(wb)
    wb.save(out_path)
    return {k: len(v) for k, v in data.items()}


# ── arithmetic audit of the markdown itself ────────────────────────────────────

def present_schedule(wb, weeks=26):
    """Presentation for the two schedule sheets — idempotent, values untouched.

    Shared by `build()` and `upgrade()` because a Gantt that prints across three pages, paints its
    week axis as if the numbers were editable levers, and clips every activity name is a chart
    nobody puts in front of a partner. Applied last, after `enforce_conventions`, which colours
    cells from what they ARE and cannot know that an axis label is not an input.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter

    if "Schedule" in wb.sheetnames:
        ws = wb["Schedule"]
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "1:1"
        ws.freeze_panes = "C2"

    if "Gantt" not in wb.sheetnames:
        return
    ws = wb["Gantt"]
    ws.column_dimensions["A"].width = 52
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:4"                    # week axis + gates + FTE repeat on every page
    ws.print_title_cols = "A:A"
    ws.freeze_panes = "B6"
    head = PatternFill("solid", fgColor=HEAD_FILL)
    for col in range(1, weeks + 2):
        c = ws.cell(row=1, column=col)
        c.fill = head
        c.font = Font(name=FONT, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
    for col in range(2, weeks + 2):                # gates read as markers, not as data
        ws.cell(row=2, column=col).font = Font(name=FONT, bold=True, color="C0392B")
        ws.cell(row=2, column=col).alignment = Alignment(horizontal="center")
    last = ws.max_row
    span = f"B6:{get_column_letter(1 + weeks)}{last}"
    # A level-of-effort line (governance, QA, review cycles) is drawn lighter than discrete work:
    # a solid 26-week bar next to solid 2-week bars reads as the biggest task on the engagement.
    # priority 1 so it beats the solid-bar rule `build()` added first — conditional formats are
    # applied in priority order, and without this the lighter LOE shade never showed
    loe = FormulaRule(formula=['AND(B6=1,INDEX(Schedule!$F:$F,ROW()-4)>0)'],
                      fill=PatternFill("solid", fgColor="9FA8DA"),
                      font=Font(color="9FA8DA", name=FONT), stopIfTrue=True)
    loe.priority = 1
    ws.conditional_formatting.add(span, loe)
    return


SHEET_ORDER = ["README", "Inputs", "RateCard", "Effort", "OverlapAudit", "Range", "Schedule",
               "Gantt", "Grades", "ClientEffort", "ScopeVariance", "Cost", "Decision",
               "Judgement"]


def order_sheets(wb):
    """Tabs in the order the model computes, so a reviewer reads it in that order.

    Appending new sheets put Schedule and Gantt after the judgement tab; and the early-return
    path of `upgrade()` skipped the reorder entirely, so whether the tabs were in a sensible
    order depended on which branch ran. One function, called from every exit.
    """
    wb._sheets = ([wb[n] for n in SHEET_ORDER if n in wb.sheetnames]
                  + [x for x in wb._sheets if x.title not in SHEET_ORDER])


def consolidate_judgement(wb):
    """Fold the legacy one-tab-per-topic judgement sheets into a single `Judgement` sheet.

    An existing workbook cannot be re-seeded (that discards the model), so the migration has to
    happen in place: carry whatever text each old tab held, then remove the tab. Idempotent.
    """
    from openpyxl.styles import Alignment as _Al
    legacy = [name for name, _p in NARRATIVE_SHEETS if name in wb.sheetnames]
    if not legacy:
        return []
    carried = {}
    for name in legacy:
        old = wb[name]
        lines = []
        for row in old.iter_rows(min_row=2, values_only=True):
            cells = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if cells:
                lines.append(" | ".join(cells) if len(cells) > 1 else cells[0])
        text = "\n".join(lines).strip()
        carried[name] = "" if text.startswith("<not yet written>") else text
        del wb[name]
    if JUDGEMENT_SHEET in wb.sheetnames:
        del wb[JUDGEMENT_SHEET]
    ws = wb.create_sheet(JUDGEMENT_SHEET)
    _widths(ws, [26, 10, 120])
    ws.append(["Topic", "State", "What it says"])
    _style(ws, 1, 3, header=True)
    r = 1
    for name, _p in NARRATIVE_SHEETS:
        r += 1
        text = carried.get(name, "")
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=f'=IF(LEN(C{r})<20,"OWED","written")')
        ws.cell(row=r, column=3, value=text or "<not yet written>")
        ws.cell(row=r, column=3).alignment = _Al(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 58 if text else 15
    ws.append([])
    ws.append(["JUDGEMENT OWED",
               f'=COUNTIF(B2:B{r},"OWED")&" of {len(NARRATIVE_SHEETS)} topics unwritten"',
               "A cost base whose basis, calibration and contingency rationale are blank is a "
               "number without an argument."])
    _style(ws, r + 2, 2, fill=DERIVED_FILL, bold=True)
    return legacy


def upgrade(xlsx_path):
    """Add sheets a newer builder knows about to an EXISTING workbook, in place.

    The workbook is the maintained artefact, so a new sheet cannot arrive by re-seeding from
    markdown — that path is for a first build only. This reads the workbook's own Effort rows
    and appends what is missing, touching nothing that already exists.
    """
    import openpyxl
    live = openpyxl.load_workbook(xlsx_path)          # formulas, not values
    consolidate_judgement(live)
    missing = [name for name in ("Schedule", "Gantt") if name not in live.sheetnames]
    if not missing:
        # nothing to add, but presentation and print setup are still worth re-applying: they are
        # idempotent and a workbook built by an older version has neither
        present_schedule(live)
        order_sheets(live)
        live.save(xlsx_path)
        return []
    eff_ws = live["Effort"]
    eff = []
    for row in eff_ws.iter_rows(min_row=2, values_only=True):
        sid = row[0]
        if not isinstance(sid, str) or not re.match(r"^S-\d+$", sid):
            continue
        eff.append((sid, row[1] or "", row[2] or 0, row[3] or 0, row[4] or 0,
                    row[7] if len(row) > 7 else ""))
    fresh_path = xlsx_path + ".upgrade.xlsx"
    build({"effort": eff, "overlap": [], "grades": [], "ratecard": [], "client": [],
           "scope": [], "certain": [], "schedule": [], "milestones": []},
          0.5, fresh_path)
    fresh = openpyxl.load_workbook(fresh_path)
    for name in missing:
        src = fresh[name]
        dst = live.create_sheet(name)
        for row in src.iter_rows():
            for cell in row:
                new_cell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
                new_cell.font, new_cell.fill = copy(cell.font), copy(cell.fill)
                new_cell.number_format, new_cell.alignment = cell.number_format, copy(cell.alignment)
        for letter, dim in src.column_dimensions.items():
            dst.column_dimensions[letter].width = dim.width
        for rng in src.conditional_formatting:
            for rule in rng.rules:
                dst.conditional_formatting.add(str(rng.sqref), rule)
        dst.freeze_panes = src.freeze_panes
    # the new sheets need their inputs and named ranges
    for nm, sheet, ref in (("term_weeks", "Inputs", None), ("days_per_week", "Inputs", None),
                           ("start_date", "Inputs", None)):
        if nm in live.defined_names:
            continue
        ws = live["Inputs"]
        r = ws.max_row + 1
        label, value, note = {
            "term_weeks": ("Term (weeks)", 26, "Contract duration. The schedule may not end after it."),
            "days_per_week": ("Working days per week", 5,
                              "Duration = effort / (people x this x utilisation)."),
            "start_date": ("Project start date", __import__("datetime").date.today(),
                           "Week 1 Monday. Every date on Schedule is derived from it."),
        }[nm]
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=note)
        _style(ws, r, 2, fill=INPUT_FILL)
        if nm == "start_date":
            ws.cell(row=r, column=2).number_format = "dd-mmm-yy"
        _named(live, nm, "Inputs", f"$B${r}")
    # Copy the new sheets' named ranges from the freshly built workbook rather than
    # re-deriving them here. They were hard-coded once, and when the Schedule gained a Span
    # column every range silently shifted one column: the FTE curve was computed from
    # Lag-as-start and Span-as-duration, produced plausible-looking numbers, and was wrong in
    # every week. One layout, one owner — `build()`.
    for name, defn in fresh.defined_names.items():
        target = str(defn.value).split("!")[0].strip("'")
        if target in missing and name not in live.defined_names:
            live.defined_names.add(defn)
    present_schedule(live)
    order_sheets(live)
    live.save(xlsx_path)
    os.remove(fresh_path)
    return missing


def export_markdown(xlsx_path, md_path):
    """Render the workbook back out as a markdown snapshot — GENERATED, never hand-edited.

    The workbook is the single maintained artefact; this exists so the estimate stays inside the
    things text gives you for free and a binary does not: `eng_lint` can read it, `git diff` can
    show what moved between two re-prices, and a reviewer can read it without opening Excel.
    Values, not formulas — so the workbook must have been recalculated first, or every derived
    cell reads back empty.

    **It overwrites, deliberately.** Git is the version history — that is most of why a text
    snapshot exists at all, and dated filenames (`estimation-2026-07-26.md`) would duplicate it,
    clutter the analysis folder and break every rule that expects one `estimation.md`. A version
    that must survive independently of git is a *freeze*, and the pack already has one place for
    that: `4_final/`, which should capture the workbook and the snapshot together.

    On `markitdown`: the `xlsx` skill names it for xlsx → markdown, and for a quick look it is
    the right tool. It is not used here for two reasons — it is a lossy preview by design ("no
    cell coordinates, don't plan edits from it"), and this snapshot is a record of record that
    has to carry the DO-NOT-EDIT banner and read post-recalculation values deterministically.
    If you want the preview instead: `pip install markitdown[xlsx]` then `markitdown file.xlsx`.
    """
    import datetime
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    fmts = openpyxl.load_workbook(xlsx_path)          # data_only loses number formats
    out = [f"# Estimation — generated from `{os.path.basename(xlsx_path)}`", "",
           "> **DO NOT EDIT THIS FILE.** It is a generated snapshot of the workbook, which is the",
           "> single maintained source. Give changes to `/engagement-os:eng-estimate`, or edit `" +
           os.path.basename(xlsx_path) + "` and invoke that skill to refresh this snapshot.",
           "> Anything typed here is lost on the next export.",
           f"> Generated {datetime.date.today().isoformat()}.", ""]
    def render(value, number_format):
        """A cell as a reader would see it in Excel, not as a float dump.

        The snapshot exists to be read and diffed, and `22.333333333333332` /
        `2026-08-03 00:00:00` / `181633.77777777778` are none of those things — the workbook
        already says what each number IS through its format, so the snapshot honours it.
        """
        if isinstance(value, (datetime.datetime, datetime.date)):
            return (value.date() if isinstance(value, datetime.datetime) else value).isoformat()
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            return "" if value is None else str(value)
        f = number_format or ""
        if "%" in f:
            return f"{value:.1%}"
        if "€" in f:
            return f"€{value:,.0f}"
        if '"x"' in f:
            return f"{value:.1f}x"
        if "." in f:
            return f"{value:.{len(f.split('.')[-1].rstrip('%€\"x'))}f}"
        if isinstance(value, float):
            return str(int(value)) if value.is_integer() else f"{value:,.2f}"
        return str(value)

    blank = 0
    for ws in wb.worksheets:
        if ws.title == "Gantt":
            # 26 columns of 1/blank is unreadable as a markdown table, and it is a PROJECTION of
            # Schedule rather than data of its own — Schedule's `Bar (wk 1 →)` column carries the
            # same shape in text, which is what a reader of the snapshot actually needs.
            out += ["## Gantt", "",
                    "Visual projection of `Schedule` (week grid, gates, weekly FTE demand) — see the",
                    "workbook. `Schedule`'s **Bar (wk 1 →)** column carries the same bars in text.", ""]
            continue
        fws = fmts[ws.title]
        rows = [[render(c.value, fws.cell(row=c.row, column=c.column).number_format)
                 for c in row] for row in ws.iter_rows()]
        rows = [r for r in rows if any(v is not None and str(v).strip() for v in r)]
        if not rows:
            continue
        out += [f"## {ws.title}", ""]
        width = max(len(r) for r in rows)
        for r in rows:
            cells = [("" if v is None else str(v)).replace("|", "\\|") for v in r]
            cells += [""] * (width - len(cells))
            if sum(1 for c in cells if c.strip()) == 1 and cells[0].strip():
                out.append(cells[0])                  # prose row, not a table row
            else:
                out.append("| " + " | ".join(cells) + " |")
                if r is rows[0]:
                    out.append("|" + "---|" * width)
        out.append("")
        blank += sum(1 for r in rows for v in r
                     if v is None and False)
    text = "\n".join(out) + "\n"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(text)
    return len(wb.worksheets)


def check(data, rho):
    """Recompute in Python and report where the markdown disagrees with its own inputs."""
    eff = data["effort"]
    missing = [k for k in TABLE_KINDS if not data[k]]
    if missing:
        print(f"  ⚠ no rows found for: {', '.join(missing)} "
              f"— is the <!--table:KIND--> marker present above each table?")
    if not eff:
        print("  no effort table — nothing to check")
        return 1
    sO, sM, sP = (sum(r[i] for r in eff) for i in (2, 3, 4))
    mean = sum((o + 4 * m + p) / 6 for _, _, o, m, p, _ in eff)
    si = math.sqrt(sum(((p - o) / 6) ** 2 for _, _, o, m, p, _ in eff))
    sc = sum((p - o) / 6 for _, _, o, m, p, _ in eff)
    sd = math.sqrt(si ** 2 + rho * (sc ** 2 - si ** 2))
    exec_band = 0.8416 * sd
    print(f"  effort rows       {len(eff)}")
    print(f"  Sum O / M / P     {sO:.0f} / {sM:.0f} / {sP:.0f}")
    print(f"  PERT mean (P50)   {mean:.1f}")
    print(f"  sigma             indep {si:.1f} · rho={rho} {sd:.1f} · full {sc:.1f}")
    print(f"  P80               {mean + exec_band:.1f}   contingency {exec_band:.1f} d")
    if data["overlap"]:
        # The audit operates on the most-likely column, so the naive baseline is Sum M + the
        # deduction — not the PERT mean, which is a different statistic.
        ded = sum(a - b for _, a, b, _ in data["overlap"])
        print(f"  overlap deduction {ded:.0f} d  (naive Sum M {sM + ded:.0f} → audited {sM:.0f})")
    if data["grades"]:
        gd = sum(g[2] for g in data["grades"])
        lab = sum(g[1] * g[2] for g in data["grades"])
        print(f"  grade days        {gd:.0f}   labour {lab:,.0f}   blended {lab / gd:,.0f}")
        if abs(gd - mean) > 0.5:
            print(f"  ⚠ grade days {gd:.0f} != P50 {mean:.1f} — re-allocate the grade mix")
    if data["scope"]:
        sv = sum(s[2] for s in data["scope"])
        print(f"  scope band        {sv:.0f} d vs execution band {exec_band:.1f} d "
              f"→ {sv / exec_band:.1f}x")
    if data["client"]:
        hrs = sum(c[2] * c[3] * c[4] for c in data["client"])
        print(f"  client hours      {hrs:.0f}  ({hrs / 7.5:.1f} client-days, "
              f"{hrs / 7.5 / mean * 100:.1f}% of our effort)")
    return 0


def find_recalc():
    """Locate the `xlsx` skill's recalc.py. We do NOT reimplement recalculation.

    openpyxl writes formulas with no cached values, so a workbook straight out of this script
    reads back as `None` to anything looking at values — and, worse, a formula that LibreOffice
    cannot evaluate bakes a literal `#NAME?` into the file we hand over. The `xlsx` skill already
    owns that problem properly: a LibreOffice macro pass that rewrites the file in place and
    returns a JSON verdict with the failing cells named. A `soffice --convert-to` call (the
    obvious hand-rolled substitute, and the one this script shipped with first) converts the file
    but reports no formula errors at all — it looks like verification and is not.
    """
    env = os.environ.get("XLSX_SKILL_DIR")
    roots = [pathlib.Path(env)] if env else []
    roots += list(pathlib.Path.home().glob(
        ".claude/plugins/marketplaces/*/skills/xlsx"))
    roots += list(pathlib.Path.home().glob(".claude/skills/xlsx"))
    for r in roots:
        c = r / "scripts" / "recalc.py"
        if c.exists():
            return c
    return None


def recalculate(path):
    """Run the xlsx skill's recalc and surface its verdict. Returns True when the file is clean."""
    script = find_recalc()
    if not script:
        print("\n  ⚠ Could not find the `xlsx` skill's scripts/recalc.py.")
        print("    Formulas are written but NOT yet evaluated, and an unevaluated workbook can")
        print("    hide a #NAME?/#REF! until the client opens it.")
        print("    → Invoke the `xlsx` skill to recalculate, or set XLSX_SKILL_DIR.")
        print("    Do NOT substitute `soffice --convert-to`: it reports no formula errors.")
        return False
    import json
    import subprocess
    r = subprocess.run([sys.executable, str(script), os.path.abspath(path)],
                       capture_output=True, text=True, cwd=str(script.parent))
    try:
        v = json.loads(r.stdout[r.stdout.index("{"):r.stdout.rindex("}") + 1])
    except Exception:
        print(f"\n  ⚠ recalc.py produced no JSON verdict:\n{r.stdout[-500:]}{r.stderr[-500:]}")
        return False
    if v.get("status") == "success":
        print(f"\n  ✓ recalculated via the xlsx skill — {v.get('total_formulas', '?')} formulas, "
              f"0 errors")
        return True
    print(f"\n  ✗ recalc reports {v.get('total_errors', '?')} formula error(s): "
          f"{v.get('error_summary')}")
    return False


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("markdown", nargs="?", help="estimation.md to seed the workbook from")
    ap.add_argument("--out", help="output .xlsx (default: alongside the markdown)")
    ap.add_argument("--blank", action="store_true", help="starter workbook, no markdown needed")
    ap.add_argument("--rho", type=float, default=0.5,
                    help="correlation assumption, 0–1 (default 0.5, the professional-services default)")
    ap.add_argument("--zero-rates", action="store_true",
                    help="emit the rate card with every rate at 0 — use when no rate card exists "
                         "and a fabricated placeholder would be worse than a visible zero")
    ap.add_argument("--to-md", action="store_true",
                    help="export the WORKBOOK back to a markdown snapshot (the normal direction "
                         "once the workbook exists — it is the maintained source)")
    ap.add_argument("--reseed", action="store_true",
                    help="rebuild the workbook FROM the markdown, discarding workbook edits")
    ap.add_argument("--no-recalc", action="store_true",
                    help="skip the xlsx-skill recalculation pass (leaves formulas unevaluated)")
    ap.add_argument("--check", action="store_true",
                    help="recompute the markdown's arithmetic and report; write nothing")
    ap.add_argument("--upgrade", action="store_true",
                    help="add sheets a newer builder knows about to an EXISTING workbook, in "
                         "place — the migration path, since re-seeding would discard the model")
    args = ap.parse_args()

    if args.upgrade:
        target = args.out or (args.markdown and
                              str(pathlib.Path(args.markdown).with_suffix(".xlsx")))
        if not target or not os.path.exists(target):
            print("--upgrade needs an existing workbook (--out path.xlsx)", file=sys.stderr)
            return 2
        added = upgrade(target)
        print(f"upgraded {os.path.basename(target)}: "
              + (", ".join(added) + " added" if added else "already current"))
        if added and not args.no_recalc:
            recalculate(target)
        return 0

    if not args.blank and not args.markdown and not (args.to_md and args.out):
        ap.error("give an estimation.md, --blank for a starter workbook, "
                 "or --to-md --out <workbook.xlsx> to export an existing one")

    data = {k: [] for k in TABLE_KINDS}
    if args.markdown:
        if not os.path.exists(args.markdown):
            print(f"ERROR: not found: {args.markdown}", file=sys.stderr)
            return 2
        with open(args.markdown, encoding="utf-8", errors="replace") as f:
            md = f.read()
        if is_generated_snapshot(md) and not args.to_md:
            # The snapshot has no `<!--table:…-->` markers, so seeding from it yields a workbook
            # of placeholder rows and zero days — and `--reseed` then writes that over the real
            # model. Refuse instead: the workbook is the maintained source, and a new sheet
            # arrives through `--upgrade`.
            print(f"ERROR: {os.path.basename(args.markdown)} is a GENERATED snapshot of the "
                  "workbook, not a seed.\n"
                  "  To add sheets a newer builder knows about:  --upgrade --out <workbook.xlsx>\n"
                  "  To refresh this snapshot from the workbook: --to-md --out <workbook.xlsx>\n"
                  "  To genuinely start over, seed from the estimation.md TEMPLATE.",
                  file=sys.stderr)
            return 2
        data = parse_tables(md)

    if args.check:
        return check(data, args.rho)

    out = args.out or (os.path.splitext(args.markdown)[0] + ".xlsx" if args.markdown
                       else "estimation.xlsx")

    if args.to_md:
        if not os.path.exists(out):
            print(f"ERROR: no workbook at {out} — seed one first", file=sys.stderr)
            return 2
        md_out = args.markdown or (os.path.splitext(out)[0] + ".md")
        # Export reads cached values. openpyxl leaves none behind, and neither does a hand edit
        # in Excel that was never reopened — so recalculate before reading, or the snapshot is
        # a page of blanks that looks like a successful export.
        if not args.no_recalc:
            recalculate(out)
        n = export_markdown(out, md_out)
        print(f"Exported {n} sheet(s) → {md_out}")
        print("The workbook stays the source of truth; this snapshot is generated.")
        return 0

    # The workbook is the maintained artefact, so seeding over an existing one destroys work.
    # This was a real defect: the docs told reviewers to edit the workbook while the builder
    # silently rebuilt it from the markdown on the next run.
    if os.path.exists(out) and not args.reseed:
        print(f"REFUSING to overwrite {out} — it is the source of truth and may hold edits\n"
              f"  · export it to markdown instead:  --to-md\n"
              f"  · or discard workbook edits on purpose:  --reseed", file=sys.stderr)
        return 5
    try:
        counts = build(data, args.rho, out, args.zero_rates)
    except ImportError:
        print("ERROR: openpyxl not installed — run: pip install openpyxl "
              "(PEP-668: add --user --break-system-packages)", file=sys.stderr)
        return 3
    print(f"Wrote {out}")
    for k, n in counts.items():
        print(f"  {k:9s} {n} row(s)" + ("   ← empty, using placeholders" if not n else ""))
    clean = True if args.no_recalc else recalculate(out)
    print("\nBlue text on yellow = input · black = formula · green = cross-sheet "
          "(the `xlsx` skill's model conventions).")
    print("The workbook is now the model — edit there, and keep the markdown in step with it.")
    return 0 if clean else 4


if __name__ == "__main__":
    sys.exit(main())
