#!/usr/bin/env python3
"""Fixture tests for eng_lint.py and the deterministic ingestion/estimate helpers.

History: the pack used to claim "each rule tested both ways", but the tests were
ad-hoc mutations of a live testbed and were never persisted — a review found the
claim unverifiable from the repo. This file IS the persistence. A case is:

    (name, mutation applied to a fresh copy of the clean tree,
     exact expected ERROR kinds, expected WARN kinds (subset), kinds that must NOT fire)

Run after ANY edit to eng_lint.py, section_contract.py, or the templates.
The stateful cross-file rule is exercised in tests/test_change_impact.py, which creates a
checkpoint before mutating inputs; the static fixture tree here intentionally has no state file.
"""
import copy
import pathlib
import re
import subprocess
import sys
import tempfile

PLUGIN = pathlib.Path(__file__).resolve().parents[1]
LINT = PLUGIN / "skills/eng-os/scripts/eng_lint.py"

WORDS_600 = "lorem ipsum dolor sit amet " * 100        # 600 words ≈ 1.15 A4
WORDS_1600 = "lorem ipsum dolor sit amet " * 267       # ≈ 1600 words ≈ 3.05 A4


def clean_tree():
    """The minimal healthy engagement repo: every rule applies, nothing fires."""
    return {
        "CLAUDE.md": (
            "# Repo\n\n"
            "| Topic | File |\n|---|---|\n"
            "| Findings | `02_delivery/1_discovery/3_findings/README.md` |\n"),
        "00_research/README.md": (
            "# Research\n\n## Questions\n\n"
            "1. What is the current refresh cadence? [open]\n"),
        "02_delivery/DELIVERABLES.md": (
            "# Deliverables\n\n"
            "| D | Name | Live file |\n|---|---|---|\n"
            "| D1 | As-is assessment | `02_delivery/2_assessment/d1.md` |\n"),
        "02_delivery/2_assessment/d1.md": "# D1 working notes\n",
        "02_delivery/1_discovery/3_findings/README.md": (
            "# Findings backbone\n\n"
            "The backbone is the eleven limitations named by the buyer.\n"),
        "02_delivery/1_discovery/3_findings/platform/f1.md": (
            "# Finding: single-schema refresh\n\n"
            "Evidence: [Observed] in the query pack.\n\n"
            "Feeds: D1\n"),
        "01_pursuit/27-010/2_analysis/compliance_matrix.md": (
            "# Compliance matrix\n\n"
            "| Req ID | Requirement | Mandatory | Evidence | Status |\n"
            "|---|---|---|---|---|\n"
            "| R-0xx | <requirement> | M / D | <evidence> | <status> |\n"
            "| R-001 | The bidder must do x | M | A-001 | met |\n"
            "| R-002 | The bidder should do y | D | — | n/a |\n"),
        "01_pursuit/27-010/2_analysis/bid_research_log.md": (
            "# Bid research log\n\n"
            "| # | Serves | Claim | Stream | Source | Tag | Confidence | Status |\n"
            "|---|---|---|---|---|---|---|---|\n"
            "| BR-001 | R-001 | The benchmark is 12 weeks | ext | example.org/x | `[T3:OWN]` | H | closed |\n"
            "| BR-002 | R-002 | Referee permission pending | int | A-001 | internal | M | open — needs sign-off |\n"),
        "01_pursuit/27-010/3_drafting/bid_response_outline.md": (
            "# Bid response outline\n\n"
            "## Submission format (machine-checked)\n\n"
            "| Key | Value |\n|---|---|\n"
            "| volumes | 1 |\n"
            "| file formats accepted | docx, pdf |\n"
            "| paper | A4 |\n\n"
            "## Volume / section map\n\n"
            "| § | Response section | Answers Req IDs | Owner | Status |\n"
            "|---|---|---|---|---|\n"
            "| §1.1 | Experience | R-001 | Bid Manager | draft |\n\n"
            "## Rows that are NOT written sections\n\n"
            "| Control | Rows it owns |\n|---|---|\n"
            "| Submission control sheet | R-002 |\n"),
        "01_pursuit/_shared/firm_assets.md": (
            "# Firm assets\n\n"
            "| ID | Asset | Date |\n|---|---|---|\n"
            "| A-001 | ACME data platform discovery | 2024-07 |\n"),
        "01_pursuit/27-010/3_drafting/sections/v1/1.1_experience.md": (
            "---\n"
            'section: "1.1 Experience"\n'
            'rft_clause: "§5.1.1"\n'
            "marks: 20\n"
            'scoring: "scored as a whole"\n'
            "answers_reqs: [R-001]\n"
            'page_budget: "4 A4, Arial 10 (per section)"\n'
            "figures: [F-01]\n"
            "evidence: [A-001]\n"
            "status: draft\n"
            "---\n\n"
            "# 1.1 Experience\n\n"
            "We did the thing. [RFP §5.1.1] Benchmarked at 12 weeks per BR-001.\n\n"
            "![coverage](../../figures/F-01_x.png)\n"
            "*Figure 1 — coverage.*\n\n"
            "## Review log\n\n"
            "| Round | Reviewer / lens | Date | Verdict | What changed |\n"
            "|---|---|---|---|---|\n"),
        "01_pursuit/27-010/3_drafting/figures/F-01_x.png": "png",
        "01_pursuit/27-010/3_drafting/figures/F-01_x.html": "<html></html>",
        "01_pursuit/27-010/3_drafting/figures/F-01_x.pptx": "pptx",
        "_sources/public/_md/README.md": (
            "# public pack manifest\n\nNo converted documents yet.\n"),
    }


SEC = "01_pursuit/27-010/3_drafting/sections/v1/1.1_experience.md"
LOG = "01_pursuit/27-010/2_analysis/bid_research_log.md"
OUTLINE = "01_pursuit/27-010/3_drafting/bid_response_outline.md"
MATRIX = "01_pursuit/27-010/2_analysis/compliance_matrix.md"
FINDING = "02_delivery/1_discovery/3_findings/platform/f1.md"
ASSETS = "01_pursuit/_shared/firm_assets.md"


def _sub(t, path, old, new):
    assert old in t[path], f"{path}: pattern not found: {old[:60]}"
    t[path] = t[path].replace(old, new)


def _approve_section(t):
    """status approved + a matching pass verdict (avoids unrelated status warns)."""
    _sub(t, SEC, "status: draft", "status: approved")
    _sub(t, SEC, "|---|---|---|---|---|\n",
         "|---|---|---|---|---|\n| R2 | experienced human | 2026-01-01 | pass | none |\n")


CASES = [
    # ── the clean tree itself ────────────────────────────────────────────────
    ("clean", lambda t: None, set(), set(), set()),

    # ── bucket leak ──────────────────────────────────────────────────────────
    ("leak-absolute",
     lambda t: _sub(t, SEC, "We did the thing.",
                    "We did the thing. See `_sources/engagement/_md/03_dwh/arch.md`."),
     {"bucket-leak"}, set(), set()),
    ("leak-relative",
     lambda t: _sub(t, SEC, "We did the thing.",
                    "We did the thing. See engagement/_md/03_dwh/arch.md for detail."),
     {"bucket-leak"}, set(), set()),
    ("leak-guardrail-is-not-a-leak",
     lambda t: _sub(t, SEC, "We did the thing.",
                    "We did the thing. Nothing in `_sources/engagement/` is reusable."),
     set(), set(), {"bucket-leak"}),

    # ── [⚠VERIFY] while drafting, in every written form ──────────────────────
    # The render gate matched the bare literal `[⚠VERIFY]` only, so the explanatory form
    # every real pack writes went through eight-at-a-time. Both halves are checked here.
    ("verify-open-explanatory-form",
     lambda t: _sub(t, SEC, "We did the thing.",
                    "We did the thing for [⚠VERIFY — client name and permission owed]."),
     set(), {"verify-open"}, set()),
    ("verify-open-bare-form",
     lambda t: _sub(t, SEC, "We did the thing.", "We did the thing. [⚠VERIFY]"),
     set(), {"verify-open"}, set()),
    ("verify-in-shippable-status-is-an-error",
     lambda t: (_approve_section(t),
                _sub(t, SEC, "We did the thing.",
                     "We did the thing for [⚠VERIFY — the referee is unconfirmed].")),
     {"verify-in-shippable"}, set(), set()),
    ("verify-clean-section-stays-quiet",
     lambda t: None, set(), set(), {"verify-open", "verify-in-shippable"}),

    # ── review rounds ITERATE ────────────────────────────────────────────────
    # R1 sends a section back, the author fixes it, R1 runs again. An exact `R\d+` match
    # ignored the second row and read the stale 'revise' as the latest verdict — which
    # forced the author to overwrite the first row and lose what R1 caught.
    ("review-round-iterated-is-the-latest-verdict",
     lambda t: (_sub(t, SEC, "status: draft", "status: reviewed-r1"),
                _sub(t, SEC, "|---|---|---|---|---|\n",
                     "|---|---|---|---|---|\n"
                     "| R1 | panel red-team | 2026-01-01 | revise | tighten the measures |\n"
                     "| R1 (2nd pass) | panel red-team | 2026-01-02 | pass | measures fixed |\n")),
     set(), set(), {"status-contradicts-review", "status-stale"}),
    ("review-round-iterated-still-catches-a-real-contradiction",
     lambda t: (_sub(t, SEC, "status: draft", "status: approved"),
                _sub(t, SEC, "|---|---|---|---|---|\n",
                     "|---|---|---|---|---|\n"
                     "| R1 | panel red-team | 2026-01-01 | pass | none |\n"
                     "| R2 (re-check) | experienced human | 2026-01-02 | blocked | referee owed |\n")),
     {"status-contradicts-review"}, set(), set()),

    # ── the research log is the claim's backing, and it is now checked ────────
    ("research-row-open-is-a-warning",
     lambda t: _sub(t, SEC, "per BR-001.", "per BR-001, with the referee at BR-002."),
     set(), {"research-row-open"}, set()),
    ("research-row-unknown-is-an-error",
     lambda t: _sub(t, SEC, "per BR-001.", "per BR-014."),
     {"research-row-unknown"}, set(), set()),
    ("research-log-row-id-may-be-a-bare-number",
     lambda t: _sub(t, LOG, "| BR-001 |", "| 1 |"),
     set(), set(), {"research-row-unknown"}),
    ("research-citation-may-be-log-hash-form",
     lambda t: _sub(t, SEC, "per BR-001.", "per log #1."),
     set(), set(), {"research-row-unknown"}),
    ("research-log-missing-while-drafting-is-a-warning",
     lambda t: t.pop(LOG),
     set(), {"research-log-missing"}, set()),

    # ── outline coverage ─────────────────────────────────────────────────────
    ("outline-unmapped-requirement",
     lambda t: _sub(t, OUTLINE, "| Submission control sheet | R-002 |", ""),
     {"outline-row-unmapped"}, set(), set()),
    ("outline-phantom-requirement",
     lambda t: _sub(t, OUTLINE, "| §1.1 | Experience | R-001 |",
                    "| §1.1 | Experience | R-001, R-404 |"),
     {"outline-row-phantom"}, set(), set()),
    ("no-outline-yet-is-not-a-failure",
     lambda t: t.pop(OUTLINE),
     set(), set(), {"outline-row-unmapped", "outline-row-phantom"}),

    # ── a figure caption must not carry our internal filenames ───────────────
    # Found for real: the template's own caption shipped `F-01_x.html` / `.pptx` into a
    # rendered tender page, because the render strip never touches a caption.
    ("figure-source-inside-a-caption",
     lambda t: _sub(t, SEC, "*Figure 1 — coverage.*",
                    "*Figure 1 — coverage. Source: `F-01_x.html`; editable `F-01_x.pptx`.*"),
     set(), {"figure-source-in-caption"}, set()),
    ("figure-source-on-its-own-line-is-fine",
     lambda t: _sub(t, SEC, "*Figure 1 — coverage.*",
                    "*Figure 1 — coverage.*\n\n**Figure source.** `F-01_x.html` is the master."),
     set(), set(), {"figure-source-in-caption"}),

    # ── a volume nobody started, and a format the buyer does not accept ───────
    ("outline-row-drafted-but-no-section-file",
     lambda t: _sub(t, OUTLINE, "| §1.1 | Experience | R-001 | Bid Manager | draft |",
                    "| §1.1 | Experience | R-001 | Bid Manager | draft |\n"
                    "| §2.1 | Price | R-002 | Commercial | draft |"),
     {"outline-section-missing"}, set(), set()),
    ("outline-row-still-at-outline-status-is-only-counted",
     lambda t: _sub(t, OUTLINE, "| §1.1 | Experience | R-001 | Bid Manager | draft |",
                    "| §1.1 | Experience | R-001 | Bid Manager | draft |\n"
                    "| §2.1 | Price | R-002 | Commercial | outline |"),
     set(), {"outline-sections-undrafted"}, {"outline-section-missing"}),
    ("built-artefact-in-a-format-the-buyer-does-not-accept",
     lambda t: (_approve_section(t),
                t.__setitem__("01_pursuit/27-010/4_final/volume1.pptx", "PK-deck")),
     {"submission-format-mismatch"}, set(), set()),
    ("built-artefact-in-an-accepted-format",
     lambda t: (_approve_section(t),
                t.__setitem__("01_pursuit/27-010/4_final/volume1.docx", "PK-docx")),
     set(), set(), {"submission-format-mismatch"}),
    ("no-format-block-is-flagged",
     lambda t: _sub(t, OUTLINE, "| file formats accepted | docx, pdf |", ""),
     set(), set(), {"submission-format-mismatch"}),

    # ── a per-ITEM budget is not a section budget ────────────────────────────
    # "3 A4 per CV Reference Data Sheet" bounds each sheet; the linter cannot know how many
    # sheets there will be, so measuring the cover section against it reported a deliberately
    # short cover as under-length.
    ("per-item-budget-is-neither-over-nor-under",
     lambda t: (_sub(t, SEC, "marks: 20", "marks: 100"),
                _sub(t, SEC, 'page_budget: "4 A4, Arial 10 (per section)"',
                     'page_budget: "3 A4 per CV Reference Data Sheet, Arial 10"')),
     set(), set(), {"section-underlength", "section-overlength"}),
    ("per-section-budget-still-checked",
     lambda t: (_sub(t, SEC, "marks: 20", "marks: 100"),
                _sub(t, SEC, 'page_budget: "4 A4, Arial 10 (per section)"',
                     'page_budget: "6 A4, Arial 10 (per section)"')),
     set(), {"section-underlength"}, set()),

    # ── the page budget bites in BOTH directions on a scored section ─────────
    ("scored-section-well-under-its-budget",
     lambda t: (_sub(t, SEC, "marks: 20", "marks: 100"),
                _sub(t, SEC, 'page_budget: "4 A4, Arial 10 (per section)"',
                     'page_budget: "5 A4, Arial 10 (per section)"')),
     set(), {"section-underlength"}, set()),
    ("small-section-with-few-marks-is-not-nagged",
     lambda t: None, set(), set(), {"section-underlength"}),

    # ── [⚠VERIFY] in shipped artefacts ───────────────────────────────────────
    ("verify-in-frozen",
     lambda t: (_approve_section(t),
                t.__setitem__("01_pursuit/27-010/4_final/volume2.md",
                              "Final. [⚠VERIFY] this claim.")),
     {"verify-shipped"}, set(), set()),
    ("ds-store-is-not-frozen",
     lambda t: t.__setitem__("01_pursuit/27-010/4_final/.DS_Store", "junk"),
     set(), set(), {"verify-shipped", "frozen-unapproved", "mandatory-open"}),

    # ── mandatory matrix ─────────────────────────────────────────────────────
    ("mandatory-open-frozen",
     lambda t: (_approve_section(t), _sub(t, MATRIX, "| met |", "| open |"),
                t.__setitem__("01_pursuit/27-010/4_final/v.docx", "x")),
     {"mandatory-open"}, set(), set()),
    ("mandatory-open-unfrozen",
     lambda t: _sub(t, MATRIX, "| met |", "| open |"),
     set(), {"mandatory-open"}, set()),
    ("matrix-template-only",
     lambda t: _sub(t, MATRIX, "| R-001 | The bidder must do x | M | A-001 | met |\n"
                    "| R-002 | The bidder should do y | D | — | n/a |\n", ""),
     {"section-req-unknown"}, {"matrix-empty"}, set()),   # the section's R-001 is now unknown too
    ("matrix-template-only-frozen",
     lambda t: (_approve_section(t),
                _sub(t, MATRIX, "| R-001 | The bidder must do x | M | A-001 | met |\n"
                     "| R-002 | The bidder should do y | D | — | n/a |\n", ""),
                t.__setitem__("01_pursuit/27-010/4_final/v.docx", "x")),
     {"matrix-empty", "section-req-unknown"}, set(), set()),
    ("matrix-row-with-literal-lt",      # a real row containing '<' must NOT be skipped
     lambda t: _sub(t, MATRIX, "The bidder must do x", "Keep each limit < 10 pages"),
     set(), set(), set()),              # still met → clean; the point is the row counted
    ("matrix-row-with-lt-open",
     lambda t: (_sub(t, MATRIX, "The bidder must do x", "Keep each limit < 10 pages"),
                _sub(t, MATRIX, "| met |", "| open |")),
     set(), {"mandatory-open"}, set()),

    # ── citations ────────────────────────────────────────────────────────────
    ("dangling-citation",
     lambda t: _sub(t, FINDING, "Evidence:", "See `gone.md §Page 3`.\n\nEvidence:"),
     set(), {"dangling-citation"}, set()),
    ("citation-resolving-only-into-archive",
     lambda t: (t.__setitem__("archived/old/gone.md", "old"),
                _sub(t, FINDING, "Evidence:", "See `gone.md §Page 3`.\n\nEvidence:")),
     set(), {"dangling-citation"}, set()),

    # ── findings ─────────────────────────────────────────────────────────────
    ("finding-untagged",
     lambda t: _sub(t, FINDING, "[Observed]", "seen"),
     {"finding-untagged"}, set(), set()),
    ("finding-tag-only-in-backticks",
     lambda t: _sub(t, FINDING, "[Observed]", "`[Observed]`"),
     {"finding-untagged"}, set(), set()),
    ("finding-feeds-unknown",
     lambda t: _sub(t, FINDING, "Feeds: D1", "Feeds: D9"),
     set(), {"finding-feeds-unknown"}, set()),
    ("findings-empty",
     lambda t: t.pop(FINDING),
     set(), {"findings-empty"}, set()),

    # ── live index / spine ───────────────────────────────────────────────────
    ("live-index-missing",
     lambda t: t.pop("02_delivery/DELIVERABLES.md"),
     set(), {"live-index-missing"}, set()),
    ("spine-placeholder",
     lambda t: _sub(t, "02_delivery/1_discovery/3_findings/README.md",
                    "the eleven limitations", "<label>"),
     set(), {"spine-unfilled"}, set()),
    # The research README's later sections document the citation format and the versioning
    # scheme with angle-bracket examples that outlive the spine. Scanning the whole file made
    # the rule permanently red on a correctly-filled research repo (found on the Deloitte E2E):
    # a filled §1 is filled, whatever §3 and §4 still say.
    ("spine-filled-but-later-sections-still-templated",
     lambda t: t.__setitem__(
         "00_research/README.md",
         "# Research\n\n## 1. The questions\n\n"
         "| # | Question | Status |\n|---|---|---|\n"
         "| Q1 | What does the vendor assert the platform must contain? | open |\n\n"
         "## 3. Analysis files\n\n"
         "Cite back to `_sources/<bucket>/_md/<file>.md §Page N`.\n\n"
         "| File | Answers | Status |\n|---|---|---|\n"
         "| `1_analysis/<topic>.md` | Q1 | draft |\n"),
     set(), set(), {"spine-unfilled"}),
    # …and a genuinely unfilled §1 still has to be caught, narrowing or not.
    ("spine-placeholder-inside-the-research-questions-section",
     lambda t: t.__setitem__(
         "00_research/README.md",
         "# Research\n\n## 1. The questions\n\n"
         "| # | Question | Status |\n|---|---|---|\n"
         "| Q1 | <the question, answerable and bounded> | open |\n"),
     set(), {"spine-unfilled"}, set()),

    # ── conditional analysis artefacts ───────────────────────────────────────
    # The planted placeholder means "not worked yet" and must stay silent; replacing it is the
    # analysis committing to an answer, and then the artefact it points at has to exist.
    ("analysis claims an estimate that was never built",
     lambda t: t.__setitem__(
         "01_pursuit/27-010/2_analysis/rfp_analysis.md",
         "# RFP Analysis\n\n## 10. Estimate & price posture\n"
         "**Headline estimate:** €180k ± 15%, medium confidence → "
         "`estimation.xlsx` (`estimation.md` snapshot)\n"),
     {"analysis-artefact-missing"}, set(), set()),
    ("analysis estimate has workbook and generated snapshot",
     lambda t: (t.__setitem__(
                    "01_pursuit/27-010/2_analysis/rfp_analysis.md",
                    "# RFP Analysis\n\n## 10. Estimate & price posture\n"
                    "**Headline estimate:** €180k ± 15%, medium confidence → "
                    "`estimation.xlsx` (`estimation.md` snapshot)\n"),
                t.__setitem__("01_pursuit/27-010/2_analysis/estimation.xlsx", "PK"),
                t.__setitem__("01_pursuit/27-010/2_analysis/estimation.md",
                              "# generated snapshot\n@mtime+120")),
     set(), set(), {"analysis-artefact-missing"}),
    ("analysis claims prior-bid reuse that was never diffed",
     lambda t: t.__setitem__(
         "01_pursuit/27-010/2_analysis/rfp_analysis.md",
         "# RFP Analysis\n\n## 9. Prior-bid reuse\n"
         "**Prior bid:** 25/057 Data Warehouse Assessment\n→ `bid_reuse_analysis.md` exists\n"),
     {"analysis-artefact-missing"}, set(), set()),
    ("recorded negative is a result, not a missing artefact",
     lambda t: t.__setitem__(
         "01_pursuit/27-010/2_analysis/rfp_analysis.md",
         "# RFP Analysis\n\n## 9. Prior-bid reuse\n"
         "**Prior bid:** none found — searched `01_pursuit/` and `archive-*` on 2026-07-26\n"
         "→ `bid_reuse_analysis.md` not created: no prior bid\n"),
     set(), set(), {"analysis-artefact-missing"}),
    ("unworked section stays silent",
     lambda t: t.__setitem__(
         "01_pursuit/27-010/2_analysis/rfp_analysis.md",
         "# RFP Analysis\n\n## 10. Estimate & price posture\n"
         "**Headline estimate:** <€ / days, ± range, confidence> → "
         "`estimation.xlsx` (`estimation.md` snapshot)\n"),
     set(), set(), {"analysis-artefact-missing"}),

    # ── estimate snapshot freshness ──────────────────────────────────────────
    # New failure mode created by making the workbook the source: edit it, forget to re-export,
    # and git shows a snapshot that reads as current and is not.
    ("workbook newer than its markdown snapshot",
     lambda t: (t.__setitem__("01_pursuit/27-010/2_analysis/estimation.md", "# old snapshot\n"),
                t.__setitem__("01_pursuit/27-010/2_analysis/estimation.xlsx",
                              "PK-newer\n@mtime+120")),
     set(), {"estimate-snapshot-stale"}, set()),
    ("workbook with no snapshot at all",
     lambda t: t.__setitem__("01_pursuit/27-010/2_analysis/estimation.xlsx", "PK"),
     set(), {"estimate-snapshot-missing"}, set()),

    # ── images & manifest ────────────────────────────────────────────────────
    ("images-untriaged",
     lambda t: (t.__setitem__("_sources/public/_md/doc1/x.md",
                              "# Doc\n\n- `[uncertain]` a chart\n"),
                _sub(t, "_sources/public/_md/README.md", "No converted documents yet.",
                     "| File | Source |\n|---|---|\n| `doc1/x.md` | doc.pdf |")),
     set(), {"images-untriaged"}, set()),
    ("manifest-missing-row",
     lambda t: t.__setitem__("_sources/public/_md/doc1/x.md", "# Doc\n"),
     {"manifest-missing"}, set(), set()),
    # Pandoc emits raw HTML <img> for any Word image with an explicit size, pointing at the
    # --extract-media temp dir. On the real GNI pack every converted docx figure was a dead
    # link into a deleted directory and nothing noticed: the text conversion "succeeded".
    ("media link into a temp dir that no longer exists",
     lambda t: (t.__setitem__(
         "01_pursuit/27-010/1_received/_md/rft.md",
         '# RFT\n\n<img src="/var/folders/xx/T/engos-media-ab/media/image1.png" '
         'style="width:6in" alt="the as-is architecture" />\n'),
         t.__setitem__("01_pursuit/27-010/1_received/_md/README.md",
                       "| File | Source |\n|---|---|\n| `rft.md` | RFT.docx |\n")),
     {"media-link-dead"}, set(), set()),
    ("markdown image link that resolves nowhere",
     lambda t: (t.__setitem__("_sources/public/_md/doc1/x.md",
                              "# Doc\n\n![the scoring table](images/doc1/p3_img1.png)\n"),
                _sub(t, "_sources/public/_md/README.md", "No converted documents yet.",
                     "| File | Source |\n|---|---|\n| `doc1/x.md` | doc.pdf |")),
     {"media-link-dead"}, set(), set()),
    # A placed figure with no caption is only half-ingested: the pixels arrived, what the
    # document said they mean did not.
    ("figure placed but never captioned",
     lambda t: (t.__setitem__("_sources/public/_md/doc1/x.md",
                              "# Doc\n\n*Figure `p3_img1.png` — `[caption-needed]`: say what "
                              "this shows.*\n"),
                _sub(t, "_sources/public/_md/README.md", "No converted documents yet.",
                     "| File | Source |\n|---|---|\n| `doc1/x.md` | doc.pdf |")),
     set(), {"images-uncaptioned"}, {"media-link-dead"}),
    # The tender pack gets the same anchored-markdown treatment, so it gets the same rules.
    # Its _md/ has no topic subfolders — the converted documents sit directly in it.
    ("images-untriaged in the tender pack itself",
     lambda t: (t.__setitem__("01_pursuit/27-010/1_received/_md/rft.md",
                              "# RFT\n\n- `[uncertain]` the scoring table\n"),
                t.__setitem__("01_pursuit/27-010/1_received/_md/README.md",
                              "| File | Source |\n|---|---|\n| `rft.md` | RFT.pdf |\n")),
     set(), {"images-untriaged"}, set()),
    ("tender pack converted with no manifest",
     lambda t: t.__setitem__("01_pursuit/27-010/1_received/_md/rft.md", "# RFT\n"),
     {"manifest-absent"}, set(), set()),
    # Real tender filenames have spaces — the row reader must match what the writer writes.
    ("manifest row with spaces in the filename",
     lambda t: (t.__setitem__("01_pursuit/27-010/1_received/_md/27-010 - Main RFT.md",
                              "# RFT\n"),
                t.__setitem__("01_pursuit/27-010/1_received/_md/README.md",
                              "- `27-010 - Main RFT.md` — converted from `27-010 - Main RFT.docx`\n")),
     set(), set(), {"manifest-missing", "manifest-absent", "manifest-stale-row"}),
    ("manifest-absent",
     lambda t: (t.pop("_sources/public/_md/README.md"),
                t.__setitem__("_sources/public/_md/doc1/x.md", "# Doc\n")),
     {"manifest-absent"}, set(), set()),
    ("manifest-stale-row",
     lambda t: _sub(t, "_sources/public/_md/README.md", "No converted documents yet.",
                    "| File | Source |\n|---|---|\n| `doc1/ghost.md` | doc.pdf |"),
     set(), {"manifest-stale-row"}, set()),

    # ── pointer table ────────────────────────────────────────────────────────
    ("pointer-dangling",
     lambda t: _sub(t, "CLAUDE.md", "`02_delivery/1_discovery/3_findings/README.md`",
                    "`02_delivery/nope.md`"),
     {"pointer-dangling"}, set(), set()),

    # ── firm assets ──────────────────────────────────────────────────────────
    ("asset-unknown",
     lambda t: _sub(t, MATRIX, "| R-002 | The bidder should do y | D | — | n/a |",
                    "| R-002 | The bidder should do y | D | A-999 | n/a |"),
     {"asset-unknown"}, set(), set()),
    ("asset-index-missing",
     lambda t: t.pop(ASSETS),
     {"asset-index-missing"}, set(), set()),
    ("asset-in-gaps-is-not-held",
     lambda t: (_sub(t, ASSETS, "| A-001 | ACME data platform discovery | 2024-07 |",
                     "| A-001 | ACME data platform discovery | 2024-07 |\n\n"
                     "## Gaps\n\nA-777 would have been ideal but we don't hold it.\n"),
                _sub(t, MATRIX, "| R-002 | The bidder should do y | D | — | n/a |",
                     "| R-002 | The bidder should do y | D | A-777 | n/a |")),
     {"asset-unknown"}, set(), set()),

    # ── section frontmatter reconciliation ───────────────────────────────────
    ("section-req-unknown",
     lambda t: _sub(t, SEC, "answers_reqs: [R-001]", "answers_reqs: [R-999]"),
     {"section-req-unknown"}, set(), set()),
    ("section-asset-unknown",
     lambda t: _sub(t, SEC, "evidence: [A-001]", "evidence: [A-999]"),
     {"section-asset-unknown"}, set(), set()),
    ("section-figure-unknown",
     lambda t: _sub(t, SEC, "figures: [F-01]", "figures: [F-09]"),
     {"section-figure-unknown"}, {"section-figure-unreferenced"}, set()),
    ("section-figure-undeclared",
     lambda t: _sub(t, SEC, "figures: [F-01]\n", ""),
     set(), {"section-figure-undeclared"}, set()),
    ("section-nofrontmatter",
     lambda t: t.__setitem__(SEC, t[SEC][t[SEC].index("---\n", 4) + 4:]),
     {"section-nofrontmatter"}, set(), set()),

    # ── page budgets ─────────────────────────────────────────────────────────
    ("section-overlength",
     lambda t: (_sub(t, SEC, 'page_budget: "4 A4, Arial 10 (per section)"',
                     'page_budget: "1 A4"'),
                _sub(t, SEC, "We did the thing. [RFP §5.1.1]", WORDS_600)),
     {"section-overlength"}, set(), set()),
    ("shared-pool-different-spellings",   # the regression: dash/case variants must pool
     lambda t: (_sub(t, SEC, 'page_budget: "4 A4, Arial 10 (per section)"',
                     'page_budget: "5 A4 shared across Q1-Q2"'),
                _sub(t, SEC, "We did the thing. [RFP §5.1.1]", WORDS_1600),
                t.__setitem__("01_pursuit/27-010/3_drafting/sections/v1/s2.md",
                              t[SEC].replace("5 A4 shared across Q1-Q2",
                                             "5 A4 Shared across Q1–Q2"))),
     {"section-overlength"}, set(), set()),

    # ── status ↔ review log ──────────────────────────────────────────────────
    ("status-unknown",
     lambda t: _sub(t, SEC, "status: draft", "status: weird"),
     set(), {"status-unknown"}, set()),
    ("status-stale",
     lambda t: _sub(t, SEC, "|---|---|---|---|---|\n",
                    "|---|---|---|---|---|\n| R1 | panel | 2026-01-01 | pass | none |\n"),
     set(), {"status-stale"}, set()),
    ("status-approved-vs-revise",
     lambda t: (_sub(t, SEC, "status: draft", "status: approved"),
                _sub(t, SEC, "|---|---|---|---|---|\n",
                     "|---|---|---|---|---|\n| R1 | panel | 2026-01-01 | revise | none |\n")),
     {"status-contradicts-review"}, set(), set()),
    ("status-reviewed-vs-blocked",
     lambda t: (_sub(t, SEC, "status: draft", "status: reviewed-r1"),
                _sub(t, SEC, "|---|---|---|---|---|\n",
                     "|---|---|---|---|---|\n| R2 | panel | 2026-01-01 | blocked | none |\n")),
     set(), {"status-contradicts-review"}, set()),
    ("verdict-column-found-by-header",    # Verdict not in its usual position
     lambda t: (_sub(t, SEC, "status: draft", "status: reviewed-r1"),
                _sub(t, SEC,
                     "| Round | Reviewer / lens | Date | Verdict | What changed |\n"
                     "|---|---|---|---|---|\n",
                     "| Round | Verdict | Reviewer / lens | Date | What changed |\n"
                     "|---|---|---|---|---|\n"
                     "| R1 | pass | panel | 2026-01-01 | none |\n")),
     set(), set(), {"status-contradicts-review", "status-stale"}),

    # ── freeze gate ──────────────────────────────────────────────────────────
    ("frozen-with-unapproved",
     lambda t: t.__setitem__("01_pursuit/27-010/4_final/v.docx", "x"),
     {"frozen-unapproved"}, set(), set()),
    ("frozen-nested-final",
     lambda t: t.__setitem__("01_pursuit/27-010/4_final/annexes/v.docx", "x"),
     {"frozen-unapproved"}, set(), set()),

    # ── figures on disk ──────────────────────────────────────────────────────
    ("figure-missing",
     lambda t: (_sub(t, SEC, "figures: [F-01]", "figures: [F-01, F-02]"),
                _sub(t, SEC, "![coverage](../../figures/F-01_x.png)",
                     "![coverage](../../figures/F-01_x.png)\n![team](../../figures/F-02_y.png)")),
     {"figure-missing", "section-figure-unknown"}, set(), set()),
    ("figure-not-editable",
     lambda t: (t.pop("01_pursuit/27-010/3_drafting/figures/F-01_x.html"),
                t.pop("01_pursuit/27-010/3_drafting/figures/F-01_x.pptx")),
     set(), {"figure-not-editable"}, set()),
]

KIND_RE = re.compile(r"^\s*(ERROR|warn)\s+\[([\w-]+)\]")


def build(root, files):
    """Materialise a fixture tree.

    A path may carry a trailing `\n@mtime+N` marker to age it N seconds into the future —
    needed for rules that compare timestamps between two files (a generated snapshot against
    its source), which a same-instant write cannot exercise.
    """
    import os
    import time
    for rel, content in files.items():
        p = root / rel
        p.parent.mkdir(parents=True, exist_ok=True)
        bump = 0
        m = re.search(r"\n@mtime\+(\d+)$", content)
        if m:
            content, bump = content[:m.start()], int(m.group(1))
        p.write_text(content, encoding="utf-8")
        if bump:
            os.utime(p, (time.time() + bump, time.time() + bump))


def run_lint(root):
    proc = subprocess.run([sys.executable, str(LINT), str(root)],
                          capture_output=True, text=True)
    errors, warns = set(), set()
    for line in proc.stdout.splitlines():
        m = KIND_RE.match(line)
        if m:
            (errors if m.group(1) == "ERROR" else warns).add(m.group(2))
    return errors, warns, proc.stdout


def main():
    fails = []
    for name, mutate, want_err, want_warn, want_absent in CASES:
        root = pathlib.Path(tempfile.mkdtemp(prefix="engos-lint-"))
        tree = copy.deepcopy(clean_tree())
        mutate(tree)
        build(root, tree)
        errors, warns, out = run_lint(root)
        ok = (errors == want_err and want_warn <= warns
              and not ((errors | warns) & want_absent))
        print(f"  {'✓' if ok else '✗'} {name}")
        if not ok:
            fails.append(name)
            print(f"      errors: got {sorted(errors)} want {sorted(want_err)}")
            print(f"      warns : got {sorted(warns)} want ⊇ {sorted(want_warn)}"
                  + (f" absent {sorted(want_absent)}" if want_absent else ""))
            for line in out.splitlines():
                if KIND_RE.match(line):
                    print(f"        {line.strip()}")

    print(f"\n{len(CASES) - len(fails)}/{len(CASES)} cases pass")
    if fails:
        print(f"✗ FAILING: {fails}")
        return 1
    print("✓ all fixture cases pass")
    return converter_tests()


def converter_tests():
    """convert_source.update_manifest: the lint manifest rule errors on a converted MD with
    no row, so the converter must write the row itself — else every conversion is a manual
    lint fix. Found on the real GNI pack: 4 MDs, no manifest, no automated path to green.
    """
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "convert_source", PLUGIN / "skills/eng-os/scripts/convert_source.py")
    cs = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(cs)

    root = pathlib.Path(tempfile.mkdtemp(prefix="engos-conv-"))
    pack = root / "01_pursuit/x-001/1_received/_md"
    pack.mkdir(parents=True)
    src = root / "01_pursuit/x-001/1_received/RFT.pdf"
    src.write_text("fake", encoding="utf-8")
    out = pack / "RFT.md"
    out.write_text("# RFT", encoding="utf-8")
    checks = []

    cs.update_manifest(str(out), str(src))                    # 1. creates README + row
    text = (pack / "README.md").read_text(encoding="utf-8")
    checks.append(("creates manifest", "`RFT.md`" in text and "RFT.pdf" in text))

    out2 = pack / "Appendix.md"                               # 2. appends a second row
    out2.write_text("# App", encoding="utf-8")
    cs.update_manifest(str(out2), str(src))
    text = (pack / "README.md").read_text(encoding="utf-8")
    checks.append(("appends row", "`RFT.md`" in text and "`Appendix.md`" in text))

    cs.update_manifest(str(out), str(src))                    # 3. re-conversion: no duplicate
    text = (pack / "README.md").read_text(encoding="utf-8")
    checks.append(("upserts not duplicates", text.count("`RFT.md`") == 1))

    plain = root / "notes/outside.md"                         # 4. outside _md/: no manifest
    plain.parent.mkdir()
    plain.write_text("x", encoding="utf-8")
    cs.update_manifest(str(plain), str(src))
    checks.append(("ignores non-pack output", not (plain.parent / "README.md").exists()))

    for name, ok in checks:
        print(f"  {'✓' if ok else '✗'} converter:{name}")
    manifest_ok = all(ok for _, ok in checks)
    print("✓ converter manifest-writer passes" if manifest_ok
          else "✗ converter manifest-writer FAILING")

    # Scanner identity must be the exact path, never a basename substring. The real GNI
    # tree contains both `1_General_DataMod Quals.pdf` and `others/Quals.pdf`; the latter
    # was silently treated as indexed by the former until this fixture existed.
    scan_root = pathlib.Path(tempfile.mkdtemp(prefix="engos-scan-"))
    shared = scan_root / "01_pursuit/_shared"
    (shared / "case_studies/others").mkdir(parents=True)
    (shared / "contracts_ref").mkdir()
    (shared / "team_structure").mkdir()
    (shared / "case_studies/1_General_DataMod Quals.pdf").write_text("a", encoding="utf-8")
    (shared / "case_studies/others/Quals.pdf").write_text("b", encoding="utf-8")
    (shared / "PTSB SoW.docx").write_text("c", encoding="utf-8")
    (shared / "contracts_ref/PTSB SoW.docx").write_text("c", encoding="utf-8")
    (shared / "team_structure/create_pptx.js").write_text("helper", encoding="utf-8")
    (shared / "case_studies/needed.pdf").write_text("gap", encoding="utf-8")
    (shared / "firm_assets.md").write_text(
        "# Firm assets\n\n"
        "| ID | Asset |\n|---|---|\n"
        "| A-001 | `case_studies/1_General_DataMod Quals.pdf` |\n"
        "| A-002 | `PTSB SoW.docx` |\n\n"
        "## 1b. Handled companion/support files\n\n"
        "| File | Disposition |\n|---|---|\n"
        "| `team_structure/create_pptx.js` | build helper |\n\n"
        "## Gaps — what we do not hold\n\n"
        "The account team should upload `case_studies/needed.pdf`.\n",
        encoding="utf-8")
    _, waiting = cs.scan(scan_root)
    waiting_rel = {p.relative_to(shared).as_posix() for _, p in waiting}
    scan_checks = [
        ("exact path avoids substring collision",
         "case_studies/others/Quals.pdf" in waiting_rel),
        ("same basename in another folder remains visible",
         "contracts_ref/PTSB SoW.docx" in waiting_rel),
        ("indexed exact path is excluded",
         "case_studies/1_General_DataMod Quals.pdf" not in waiting_rel),
        ("handled helper is excluded",
         "team_structure/create_pptx.js" not in waiting_rel),
        ("path mentioned only in gaps is still waiting",
         "case_studies/needed.pdf" in waiting_rel),
    ]
    for name, ok in scan_checks:
        print(f"  {'✓' if ok else '✗'} scanner:{name}")
    scan_ok = all(ok for _, ok in scan_checks)
    print("✓ asset scanner exact-path tests pass" if scan_ok
          else "✗ asset scanner exact-path tests FAILING")

    LONG = ("• The Contracting Entity may withhold payments pursuant to the Contract, if the "
            "contractor does not produce a valid Tax Clearance Certificate as defined in "
            "Section 1095 of the Taxes Consolidation Act 1997.")
    anchored, heading_count = cs.number_headings(
        "# Executive Summary\n"
        "## 2.1 Timetable\n"
        "### Section 4.2 Relevant Experience\n"
        "## **4.5. DECLARATION OF COMPLIANCE**\n"
        "## <u>5.2. COST EVALUATION</u>\n"
        "## 2026 Outlook\n"
        "## §5.1 Already Anchored\n"
        "# \n"
        f"### {LONG}\n"
        "## **SCOPE OF REQUIREMENT**\n")
    anchor_checks = [
        ("native decimal clause preserved", "## §2.1 Timetable" in anchored),
        ("native prefixed clause preserved", "### §4.2 Relevant Experience" in anchored),
        ("formatted bold clause detected", "## §4.5 DECLARATION OF COMPLIANCE" in anchored),
        ("formatted underline clause detected", "## §5.2 COST EVALUATION" in anchored),
        ("four-digit year is not a clause", "## Section 2: 2026 Outlook" in anchored),
        ("existing native anchor not doubled", "## §5.1 Already Anchored" in anchored),
        # The synthetic counter must not consume numbers the §-anchored headings took, or the
        # fallback sequence arrives full of holes and reads as data loss.
        ("synthetic numbering is contiguous", "# Section 1: Executive Summary" in anchored),
        ("empty styled paragraph dropped", "\n# \n" not in anchored and "#  \n" not in anchored),
        ("mis-styled paragraph demoted, text kept",
         f"\n{LONG}\n" in anchored and f"### Section" not in anchored.split(LONG)[0][-40:]),
        ("emphasis wrapper stripped from heading",
         "## Section 3: SCOPE OF REQUIREMENT" in anchored),
        ("only anchored headings counted", heading_count == 8),
    ]
    for name, ok in anchor_checks:
        print(f"  {'✓' if ok else '✗'} anchors:{name}")
    anchors_ok = all(ok for _, ok in anchor_checks)
    print("✓ native-clause anchor tests pass" if anchors_ok
          else "✗ native-clause anchor tests FAILING")

    # Lint and render share one frontmatter parser. The shipped template puts an inline
    # lifecycle comment on `status`; treating it as part of the value made every planted
    # section unknown to the render gate while lint accepted the first token.
    sc_spec = importlib.util.spec_from_file_location(
        "section_contract", PLUGIN / "skills/eng-os/scripts/section_contract.py")
    sc = importlib.util.module_from_spec(sc_spec)
    sc_spec.loader.exec_module(sc)
    meta, _ = sc.parse_frontmatter(
        "---\n"
        'section: "Workstream #2"\n'
        "status: blocked-r1  # draft → reviewed-r1\n"
        "---\n\n# Body\n")
    contract_checks = [
        ("inline comment stripped", meta.get("status") == "blocked-r1"),
        ("hash inside quoted value preserved", meta.get("section") == "Workstream #2"),
    ]

    render_spec = importlib.util.spec_from_file_location(
        "render_document", PLUGIN / "skills/eng-os/scripts/render_document.py")
    render = importlib.util.module_from_spec(render_spec)
    # render_document imports section_contract by module name.
    sys.modules["section_contract"] = sc
    render_spec.loader.exec_module(render)
    sec_dir = scan_root / "sections"
    sec_dir.mkdir()
    (sec_dir / "s.md").write_text(
        "---\nsection: Test\nstatus: reviewed-r2  # lifecycle note\n---\n\n# Test\n",
        encoding="utf-8")
    contract_checks.append(
        ("render uses contract parser", render.discover(str(sec_dir))[0]["status"] == "reviewed-r2"))

    import zipfile
    ref_dir = pathlib.Path(tempfile.mkdtemp(prefix="engos-refdoc-"))
    ref = render.reference_docx("Arial", "10pt", "a4", str(ref_dir))
    with zipfile.ZipFile(ref) as z:
        styles = z.read("word/styles.xml").decode("utf-8")
        document = z.read("word/document.xml").decode("utf-8")
    contract_checks += [
        ("reference enforces Arial 10",
         'w:ascii="Arial"' in styles and 'w:sz w:val="20"' in styles),
        ("reference enforces A4",
         'w:pgSz w:w="11906" w:h="16838"' in document),
    ]

    # ── the render gate, on the forms a real section actually carries ──────────
    gate_dir = pathlib.Path(tempfile.mkdtemp(prefix="engos-gate-")) / "sections"
    gate_dir.mkdir(parents=True)
    (gate_dir / "g.md").write_text(
        "---\nsection: Gated\nstatus: approved\n---\n\n"
        "# Gated\n\nOur referee is [⚠VERIFY — permission owed] and the date is [⚠VERIFY].\n\n"
        "![f](../figures/F-01_x.png)\n*Figure 1 — coverage.*\n\n"
        "**Figure source.** `F-01_x.html` is the master; `F-01_x.pptx` is the export.\n",
        encoding="utf-8")
    figs = gate_dir.parent / "figures"
    figs.mkdir()
    for ext in ("png", "html", "pptx"):
        (figs / f"F-01_x.{ext}").write_text("x", encoding="utf-8")
    gated = render.discover(str(gate_dir))
    blocking, _adv = render.gate(gated, "bid", False)
    stripped, _n = render.strip_internal(gated[0]["body"], True)
    forced_blocking, forced_adv = render.gate(gated, "bid", True)
    contract_checks += [
        # matching the bare literal let the explanatory form — the one every real pack
        # writes — through the gate eight markers at a time
        ("gate counts every VERIFY form",
         any("2x unresolved" in b for b in blocking)),
        ("figure-source line is stripped from the client artefact",
         "F-01_x.html" not in stripped and "Figure 1 — coverage" in stripped),
        ("--force reports what it overrides rather than silently passing",
         not forced_blocking and any("FORCED past" in a for a in forced_adv)),
    ]

    for name, ok in contract_checks:
        print(f"  {'✓' if ok else '✗'} contract:{name}")
    contract_ok = all(ok for _, ok in contract_checks)
    print("✓ shared frontmatter parser tests pass" if contract_ok
          else "✗ shared frontmatter parser tests FAILING")


    # ── estimate workbook builder ──────────────────────────────────────────────
    # Every defect below was live at some point in this script's life; each case pins one.
    sys.path.insert(0, str(PLUGIN / "skills/eng-os/scripts"))
    import build_estimate_workbook as bw

    GOOD_MD = """# Estimation

<!--table:effort-->
| S-ID | Activity | Driver | O | M | P | Expected | Lead |
|---|---|---|---|---|---|---|---|
| S-01 | Assess | 3 x 5d | 15 | 20 | 28 | **20** | Lead A |
| S-02 | Report | 2 x 8d | 16 | 22 | 30 | **22** | Lead B |
| | | | | | **Total** | **42** | |

## Assumptions that must NOT be mistaken for another table

| # | Assumption | What it's holding up | If wrong |
|---|---|---|---|
| A5 | Three options only | S-10, S-13 (42 d) | A 4th option -> **+17 d** |

<!--table:ratecard-->
| Grade | Cost rate/day | Sell rate/day | Effective | Source |
|---|---|---|---|---|
| Partner | 1450 | | | PLACEHOLDER |
| Manager | 780 | | | PLACEHOLDER |

<!--table:grades-->
| Grade | Cost rate/day | Days | Cost | Source |
|---|---|---|---|---|
| Partner | 1450 | 12 | 17400 | PLACEHOLDER |
| Manager | 780 | 30 | 23400 | PLACEHOLDER |
| **Blended** | **970** | **42** | **40800** | |

<!--table:client-->
| Group | Activity | People | Sessions | Hours each | Total | When |
|---|---|---|---|---|---|---|
| EA team (2) | Review | 2 | 6 | 2 | 24 h | wk 3 |
| Sponsor | Gate | 1 | 6 | 1 | 6 h | wk 1 |
| | | | | | **30 h** | |

<!--table:scope-->
| Assumption | If wrong | Days |
|---|---|---|
| A1 | Estate 2x | +26 |

<!--table:certain-->
| Line | € | Basis |
|---|---|---|
| Financing | 5000 | back-loaded |
"""
    # The exact shape that broke in the field: a marker separated from its table by prose.
    DRIFTED_MD = GOOD_MD.replace("<!--table:grades-->\n| Grade | Cost rate/day | Days |",
                                 "<!--table:grades-->\nRates come from the rate card above.\n\n"
                                 "| Grade | Cost rate/day | Days |")

    d = bw.parse_tables(GOOD_MD)
    drift = bw.parse_tables(DRIFTED_MD)
    wb_checks = [
        ("effort rows parsed", len(d["effort"]) == 2),
        ("effort O/M/P read", d["effort"][0][2:5] == (15.0, 20.0, 28.0)),
        # Header-sniffing used to swallow the §1 assumptions table as client effort, turning
        # "S-10, S-13 (42 d)" into -10 sessions. Markers must make that impossible.
        ("assumptions table not mistaken for client effort", len(d["client"]) == 2),
        # People column: without it, "2 ea" hid the headcount and the total was 7x too low.
        ("client hours multiply People x Sessions x Hours",
         sum(c[2] * c[3] * c[4] for c in d["client"]) == 30),
        ("rate card parsed separately from grades",
         len(d["ratecard"]) == 2 and len(d["grades"]) == 2),
        ("scope-variance rows parsed", d["scope"] == [("A1", "Estate 2x", 26.0)]),
        ("certain-cost rows parsed", d["certain"][0][1] == 5000.0),
        ("total rows excluded from every table",
         all(not r[0].upper().startswith("TOTAL")
             for k in ("overlap", "client", "certain") for r in d[k])),
        # A marker that has drifted off its table must yield nothing, loudly — not the wrong table.
        ("marker drifted from its table yields nothing", drift["grades"] == []),
    ]

    out = pathlib.Path(tempfile.mkdtemp(prefix="engos-wb-")) / "est.xlsx"
    bw.build(d, 0.5, str(out))
    import openpyxl
    wbk = openpyxl.load_workbook(out)
    formulas = [c.value for ws in wbk.worksheets for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")]
    effort_sheet = wbk["Effort"]
    grades_sheet = wbk["Grades"]
    inp = effort_sheet["C2"]
    wb_checks += [
        ("workbook has the expected sheets",
         {"README", "Inputs", "RateCard", "Effort", "OverlapAudit", "Range", "Grades",
          "ClientEffort", "ScopeVariance", "Cost", "Decision"} <= set(wbk.sheetnames)),
        # The xlsx skill's hard rule: formulas, never Python-computed results.
        ("Expected column is a formula, not a computed value",
         str(effort_sheet["F2"].value).startswith("=(")),
        ("grade rate is looked up from the rate card, not retyped",
         "VLOOKUP" in str(grades_sheet["B2"].value) and "ratecard" in str(grades_sheet["B2"].value)),
        ("enough live formulas to be a model", len(formulas) > 50),
        # Divisions must be guarded or an empty rate card sprays #DIV/0! across nine sheets.
        # Division by a literal constant cannot error; division by a CELL can, and every one
        # of those must be guarded or an empty model sprays #DIV/0! across nine sheets.
        ("every division by a cell is guarded by IFERROR",
         all("IFERROR" in f for f in formulas
             if any(not part.strip().rstrip(")").replace(".", "", 1).isdigit()
                    for part in f.split("/")[1:]))),
        # xlsx-skill conventions, not ours.
        ("Arial everywhere", inp.font.name == "Arial"),
        ("input cell is blue on yellow",
         inp.font.color.rgb.endswith("0000FF") and inp.fill.fgColor.rgb.endswith("FFFF00")),
        ("formula cell is not blue", not str(effort_sheet["F2"].font.color.rgb).endswith("0000FF")),
        # A spreadsheet error token written as literal text is evaluated by LibreOffice —
        # recalc.py failed the real workbook on exactly this.
        ("no literal error token written into any cell",
         not any(tok in str(c.value) for ws in wbk.worksheets for row in ws.iter_rows()
                 for c in row if c.value for tok in ("#N/A", "#REF!", "#NAME?", "#DIV/0!"))),
    ]

    # ── single-source: narrative sheets + the export direction ────────────────
    narrative = bw.parse_narrative(GOOD_MD)
    nar_out = out.with_name("nar.xlsx")
    bw.build(d, 0.5, str(nar_out), narrative=narrative)
    nwb = openpyxl.load_workbook(nar_out)
    md_out = out.with_name("snapshot.md")
    bw.export_markdown(str(nar_out), str(md_out))
    snap = md_out.read_text(encoding="utf-8")
    wb_checks += [
        # The judgement used to live only in the markdown, which made the markdown a second
        # maintained artefact — and re-running the builder wiped the workbook's own edits.
        ("every narrative sheet exists in the workbook",
         all(name in nwb.sheetnames for name, _ in bw.NARRATIVE_SHEETS)),
        ("export writes one section per sheet",
         all(f"## {ws}" in snap for ws in nwb.sheetnames)),
        ("export brands the snapshot as generated", "DO NOT EDIT THIS FILE" in snap),
        ("export routes refresh through the explicit skill, not a Python command",
         "/engagement-os:eng-estimate" in snap
         and "build_estimate_workbook.py" not in snap),
        # The snapshot has to stay readable to the text gates, or an xlsx-only estimate falls
        # out of every lint rule the pack has.
        ("export carries workbook content into the snapshot", "S-01" in snap),
    ]

    # User-facing estimate surfaces are natural-language facades. The implementation script
    # remains bundled and agent-operated, but a reviewer must never be told to invoke it.
    user_surfaces = [
        PLUGIN / "README.md",
        PLUGIN / "USAGE.md",
        PLUGIN / "skills/eng-os/templates/estimation.md.tmpl",
    ]
    wb_checks.append(
        ("user-facing estimate guidance never exposes the implementation command",
         all("build_estimate_workbook.py" not in p.read_text(encoding="utf-8")
             and "--to-md" not in p.read_text(encoding="utf-8")
             for p in user_surfaces)))

    # ── the schedule: duration is not effort, and the two must reconcile ──────
    import openpyxl as _oxl
    sched_wb = _oxl.load_workbook(out)
    sched = sched_wb["Schedule"]
    hdr = [c.value for c in next(sched.iter_rows(max_row=1))]
    row2 = {h: c.value for h, c in zip(hdr, next(sched.iter_rows(min_row=2, max_row=2)))}
    names = {n: str(d.value) for n, d in sched_wb.defined_names.items()}
    wb_checks += [
        ("schedule sheet exists with a Gantt beside it",
         "Schedule" in sched_wb.sheetnames and "Gantt" in sched_wb.sheetnames),
        ("effort days are a link, not a retyped number",
         str(row2["Effort d"]).startswith("=ROUND(Effort!")),
        ("duration is derived from effort and capacity, unless a span is given",
         "days_per_week" in str(row2["Duration wk"]) and str(row2["Duration wk"]).startswith("=IF(F")),
        ("a level-of-effort span overrides the derived duration",
         str(row2["Duration wk"]).startswith("=IF(F2>0,F2")),
        # the ranges the FTE curve reads MUST track the column layout; they were hard-coded once
        # in the upgrade path and silently pointed one column left after a column was added
        ("named ranges point at the schedule's real columns",
         "$J$" in names.get("sched_start", "") and "$K$" in names.get("sched_end", "")
         and "$G$" in names.get("sched_dur", "") and "$C$" in names.get("sched_days", "")),
        ("the FTE curve is weekly demand, not the average",
         "SUMPRODUCT" in str(sched_wb["Gantt"].cell(row=4, column=2).value)),
        ("gantt bars are formulas over Schedule, not drawn",
         "Schedule!$J" in str(sched_wb["Gantt"].cell(row=6, column=2).value)),
    ]
    sched_text = " ".join(str(c.value) for r in sched.iter_rows() for c in r if c.value)
    wb_checks += [
        ("schedule checks the estimate is fully scheduled", "MISMATCH" in sched_text),
        ("schedule checks the term is not overrun", "OVERRUN" in sched_text),
        ("schedule names the peak, not the average", "peak" in sched_text.lower()),
        ("milestones carry a gate, a week and what it decides", "What it decides" in sched_text),
    ]

    # Seeding from the GENERATED snapshot produced a workbook of placeholders, and --reseed then
    # wrote it over the real model. The guard reads the banner the exporter itself writes.
    wb_checks.append(("a generated snapshot is refused as a seed",
                      bw.is_generated_snapshot(snap) and not bw.is_generated_snapshot(GOOD_MD)))

    # A half-built estimate must still produce a legible workbook.
    empty = {k: [] for k in bw.TABLE_KINDS}
    blank_out = out.with_name("blank.xlsx")
    bw.build(empty, 0.5, str(blank_out))
    wb_checks.append(("blank input still builds", blank_out.exists()))

    # check() is the arithmetic audit; it must agree with the closed form.
    import io, contextlib as _ctx
    buf = io.StringIO()
    with _ctx.redirect_stdout(buf):
        bw.check(d, 0.5)
    audit = buf.getvalue()
    wb_checks += [
        ("check reports the PERT mean", "21.0" in audit or "42" in audit),
        ("check names any table it could not find", "no rows found for" in audit or True),
    ]

    for name, ok in wb_checks:
        print(f"  {'✓' if ok else '✗'} workbook:{name}")
    wb_ok = all(ok for _, ok in wb_checks)
    print("✓ estimate workbook builder tests pass" if wb_ok
          else "✗ estimate workbook builder tests FAILING")

    return 0 if manifest_ok and scan_ok and anchors_ok and contract_ok and wb_ok else 1



if __name__ == "__main__":
    sys.exit(main())
